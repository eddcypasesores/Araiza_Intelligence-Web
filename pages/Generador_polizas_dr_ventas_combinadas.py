"""Generador para Pólizas Dr Ventas Combinadas integrado al hub principal."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any
from urllib.parse import urlencode

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from core.auth import ensure_session_from_token, auth_query_params
from core.custom_nav import handle_logout_request, render_brand_logout_nav


def _get_params() -> dict[str, str]:
    try:
        raw = st.query_params
    except Exception:
        raw = st.experimental_get_query_params()

    flattened: dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(value, list):
            value = value[-1] if value else None
        if value is None:
            continue
        flattened[key] = str(value)
    return flattened


def _handle_pending_navigation() -> None:
    params = _get_params()
    goto = params.pop("goto", None)
    if not goto:
        return
    try:
        st.query_params.clear()
        if params:
            st.query_params.update(params)
    except Exception:
        st.experimental_set_query_params(**params)
    try:
        st.switch_page(goto)
    except Exception:
        st.stop()
    st.stop()


def _back_href() -> str:
    params = {"goto": "pages/generador_polizas.py"}
    params.update(auth_query_params())
    query = urlencode(params, doseq=False)
    return f"?{query}"


st.set_page_config(page_title="Pólizas Dr Ventas Combinadas", layout="wide", initial_sidebar_state="collapsed")
ensure_session_from_token()
handle_logout_request()
_handle_pending_navigation()

PAGE_CSS = """
<style>
#MainMenu, header[data-testid="stHeader"], footer, div[data-testid="stToolbar"], [data-testid="stSidebar"] {
  display:none !important;
}
.block-container {
  padding-top: 120px !important;
  max-width: 900px !important;
}
body, [data-testid="stAppViewContainer"] {
  background:#f5f6fb !important;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

render_brand_logout_nav(
    "pages/generador_polizas.py",
    brand="Generador de pólizas",
    action_label="Atrás",
    action_href=_back_href(),
)


def extract_day(value: Any):
    if isinstance(value, datetime):
        return value.day
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
            return dt.day
        except Exception:
            try:
                return int(value)
            except Exception:
                return value
    if isinstance(value, str):
        fmts = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"]
        for f in fmts:
            try:
                return datetime.strptime(value.strip(), f).day
            except Exception:
                pass
        return value
    return value


def fila_vacia(valores):
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in valores)


def write_block(
    ws_out,
    base_row,
    src_A,
    src_B,
    src_C,
    src_D,
    src_E,
    src_F,
    consecutivo,
    cod_ingresos="4101-001-000000",
    cod_iva_tras="2104-001-000000",
    cod_ingresos_alt="4101-002-000",
):
    r = base_row

    ws_out.cell(row=r, column=1, value="Dr")
    ws_out.cell(row=r, column=2, value=int(consecutivo))
    ws_out.cell(row=r, column=3, value=f"Provisión Fact. {'' if src_D is None else src_D}")
    ws_out.cell(row=r, column=4, value=extract_day(src_A))

    ws_out.cell(row=r + 1, column=2, value=src_B)
    ws_out.cell(row=r + 1, column=3, value=0)
    ws_out.cell(row=r + 1, column=4, value=src_C)
    ws_out.cell(row=r + 1, column=5, value=1)
    ws_out.cell(row=r + 1, column=6, value=src_F)
    ws_out.cell(row=r + 1, column=7, value=0)

    ws_out.cell(row=r + 2, column=2, value=cod_ingresos)
    ws_out.cell(row=r + 2, column=3, value=0)
    ws_out.cell(row=r + 2, column=4, value="Ingresos al 16%")
    ws_out.cell(row=r + 2, column=5, value=1)
    ws_out.cell(row=r + 2, column=6, value=0)
    ws_out.cell(row=r + 2, column=7, value=f"=G{r+4}/0.16")
    ws_out.cell(row=r + 2, column=7).number_format = "0.00"

    ws_out.cell(row=r + 3, column=2, value=cod_ingresos_alt)
    ws_out.cell(row=r + 3, column=3, value=0)
    ws_out.cell(row=r + 3, column=4, value="Ingresos al 0%")
    ws_out.cell(row=r + 3, column=5, value=1)
    ws_out.cell(row=r + 3, column=6, value=0)
    ws_out.cell(row=r + 3, column=7, value=f"=F{r+1}-G{r+2}-G{r+4}")
    ws_out.cell(row=r + 3, column=7).number_format = "0.00"

    ws_out.cell(row=r + 4, column=2, value=cod_iva_tras)
    ws_out.cell(row=r + 4, column=3, value=0)
    ws_out.cell(row=r + 4, column=4, value="IVA trasladado no cobrado")
    ws_out.cell(row=r + 4, column=5, value=1)
    ws_out.cell(row=r + 4, column=6, value=0)
    ws_out.cell(row=r + 4, column=7, value=src_E)
    ws_out.cell(row=r + 4, column=7).number_format = "0.00"

    ws_out.cell(row=r + 5, column=2, value="FIN_PARTIDAS")


def build_stacked_output(
    src_bytes: bytes,
    consecutivo_inicial: int,
    cod_ingresos="4101-001-000000",
    cod_iva_tras="2104-001-000000",
    cod_ingresos_alt="4101-002-000",
) -> bytes:
    wb_src = load_workbook(filename=BytesIO(src_bytes), data_only=True)
    ws_src = wb_src.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Hoja1"

    current = 3
    max_row = ws_src.max_row
    generadas = 0
    consecutivo = int(consecutivo_inicial)

    for r in range(2, max_row + 1):
        A = ws_src.cell(row=r, column=1).value
        B = ws_src.cell(row=r, column=2).value
        C = ws_src.cell(row=r, column=3).value
        D = ws_src.cell(row=r, column=4).value
        E = ws_src.cell(row=r, column=5).value
        F = ws_src.cell(row=r, column=6).value

        if fila_vacia((A, B, C, D, E, F)):
            continue

        write_block(
            ws_out,
            current,
            A,
            B,
            C,
            D,
            E,
            F,
            consecutivo,
            cod_ingresos,
            cod_iva_tras,
            cod_ingresos_alt,
        )
        current += 6
        consecutivo += 1
        generadas += 1

    if generadas == 0:
        raise ValueError("No se encontraron filas con datos desde la fila 2 en las columnas A–F.")

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.getvalue()


def render_tab(
    tab_label,
    key_prefix,
    cod_ingresos_override=None,
    cod_iva_tras_override=None,
    cod_ingresos_alt_override=None,
):
    st.subheader(f"Proceso {tab_label}")
    uploaded = st.file_uploader(
        "Selecciona un archivo Excel (.xlsx)",
        type=["xlsx"],
        key=f"{key_prefix}_uploader",
    )
    consec_ini = st.number_input(
        "Consecutivo inicial (B3 del primer proceso)",
        min_value=0,
        value=1,
        step=1,
        key=f"{key_prefix}_consec",
    )

    if uploaded:
        if st.button("Generar archivo apilado", key=f"{key_prefix}_btn"):
            try:
                data = build_stacked_output(
                    uploaded.getvalue(),
                    consecutivo_inicial=consec_ini,
                    cod_ingresos=cod_ingresos_override or "4101-001-000000",
                    cod_iva_tras=cod_iva_tras_override or "2104-001-000000",
                    cod_ingresos_alt=cod_ingresos_alt_override or "4101-002-000",
                )
                st.success(f"Archivo generado con todos los procesos apilados ({tab_label}).")
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=data,
                    file_name=f"Salidas_Apiladas_{tab_label}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"{key_prefix}_dl",
                )
            except Exception as e:
                st.error(f"Ocurrió un error: {e}")
        with st.expander("Vista rápida (fila 2)"):
            try:
                wb_tmp = load_workbook(filename=BytesIO(uploaded.getvalue()), data_only=True)
                ws_tmp = wb_tmp.active
                st.write(
                    {
                        "A2 (para día)": ws_tmp["A2"].value,
                        "B2 → B4": ws_tmp["B2"].value,
                        "C2 → D4": ws_tmp["C2"].value,
                        "D2 → C3": ws_tmp["D2"].value,
                        "F2 → F4": ws_tmp["F2"].value,
                        "E2 → G7": ws_tmp["E2"].value,
                        "Última fila detectada": ws_tmp.max_row,
                    }
                )
            except Exception as e:
                st.warning(f"No se pudieron mostrar detalles: {e}")
    else:
        st.info("Esperando un archivo .xlsx…")


st.title("Generador de Pólizas Dr Ventas Combinadas")

tabs = st.tabs(["4-3-3", "4-3-4", "4-3-6"])

with tabs[0]:
    render_tab(
        "4-3-3",
        "p433",
        cod_ingresos_override="4101-001-000",
        cod_iva_tras_override="2104-001-000",
        cod_ingresos_alt_override="4101-002-000",
    )

with tabs[1]:
    render_tab(
        "4-3-4",
        "p434",
        cod_ingresos_override="4101-001-0000",
        cod_iva_tras_override="2104-001-0000",
        cod_ingresos_alt_override="4101-002-0000",
    )

with tabs[2]:
    render_tab(
        "4-3-6",
        "p436",
        cod_ingresos_override="4101-001-000000",
        cod_iva_tras_override="2104-001-000000",
        cod_ingresos_alt_override="4101-002-000000",
    )
