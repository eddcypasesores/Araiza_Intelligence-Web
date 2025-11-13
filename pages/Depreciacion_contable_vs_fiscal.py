# pages/Depreciacion_contable_vs_fiscal.py
from __future__ import annotations
import io
from pathlib import Path
from datetime import datetime
from typing import List

import pandas as pd
import streamlit as st
from core.theme import apply_theme
from core.custom_nav import render_brand_logout_nav
from openpyxl import load_workbook

# âââââââââââââââââââââââââââââ
# Config
# âââââââââââââââââââââââââââââ
st.set_page_config(
    page_title="CÃ©dula de deducciÃ³n anual",
    layout="wide",
    initial_sidebar_state="collapsed",
)
apply_theme()

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
TPL_CEDULA = (
    ASSETS_DIR / "Cedula_Deduccion_Anual_v8.xlsx"
    if (ASSETS_DIR / "Cedula_Deduccion_Anual_v8.xlsx").exists()
    else Path("Cedula_Deduccion_Anual_v8.xlsx")
)

# âââââââââââââââââââââââââââââ
# RedirecciÃ³n del botÃ³n "AtrÃ¡s"
# âââââââââââââââââââââââââââââ
# Si la URL trae ?back=1, manda a Cedula_Impuestos_inicio
try:
    qp = dict(st.query_params)  # Streamlit >= 1.32
except Exception:
    qp = st.experimental_get_query_params()

if ("back" in qp) and (qp.get("back") in (["1"], "1", 1, True)):
    try:
        st.switch_page("pages/Cedula_Impuestos_inicio.py")
    except Exception:
        # En algunos entornos el slug puede variar; si falla, no rompemos la app.
        pass

# âââââââââââââââââââââââââââââ
# CSS GLOBAL: ocultar totalmente UI nativa de Streamlit + navbar
# âââââââââââââââââââââââââââââ
GLOBAL_CSS = """
<style>
  [data-testid="stSidebar"], [data-testid="stSidebarNav"],
  header[data-testid="stHeader"], div[data-testid="stToolbar"],
  #MainMenu, #stDecoration, footer, [data-testid="stStatusWidget"],
  .viewerBadge_container__1QSob, .stDeployButton {
    display: none !important;
    visibility: hidden !important;
  }
  section.main > div.block-container { padding-top: 0.5rem !important; }
  @media (min-width: 0px){
    section[data-testid="stSidebar"] + div[role="main"] { margin-left: 0 !important; }
  }
  .nav-spacer{ height: 90px; }
</style>
"""
st.markdown(GLOBAL_CSS, unsafe_allow_html=True)

# âââââââââââââââââââââââââââââ
# Navbar (link que dispara ?back=1)
# âââââââââââââââââââââââââââââ
render_brand_logout_nav(
    "pages/Cedula_Impuestos_inicio.py",
    brand="C?dula deducci?n anual",
    action_label="Atrás",
    action_href="?back=1",
)
st.markdown('<div class="nav-spacer"></div>', unsafe_allow_html=True)

# âââââââââââââââââââââââââââââ
# Helpers
# âââââââââââââââââââââââââââââ
def _to_decimal_percent(val) -> float:
    """Acepta 10, '10', '10%' y los convierte a 0.10. Si ya es 0.10 lo deja igual."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, str):
        v = val.strip().replace("%", "").replace(",", ".")
        try:
            num = float(v)
        except Exception:
            return 0.0
    else:
        try:
            num = float(val)
        except Exception:
            return 0.0
    return num / 100.0 if num > 1.0 else num

def _load_catalog_options_from_template() -> List[str]:
    """Usa exactamente los nombres del catÃ¡logo para garantizar el match del % en la cÃ©dula."""
    if not TPL_CEDULA.exists():
        return []
    df_cat = pd.read_excel(TPL_CEDULA, sheet_name="Catalogo_Clasificacion")
    if "ClasificaciÃ³n" in df_cat.columns:
        return [str(x) for x in df_cat["ClasificaciÃ³n"].dropna().tolist()]
    return []

# âââââââââââââââââââââââââââââ
# Interfaz sin datos precargados
# âââââââââââââââââââââââââââââ
st.title("CÃ©dula de deducciÃ³n anual")
st.caption("Agrega filas con â+ Add rowâ. Si dejas vacÃ­o el % fiscal, la cÃ©dula lo toma del catÃ¡logo.")

CLASIFICACIONES = _load_catalog_options_from_template()

columns_order = [
    "ClasificaciÃ³n para el llenado de la declaraciÃ³n",
    "Fecha de adquisiciÃ³n",
    "DescripciÃ³n",
    "MOI",
    "Limite de la deducciÃ³n",
    "DepreciaciÃ³n acumulada",
    "Meses de Utilizacion",
    "% de deducciÃ³n fiscal (capturar como factor ejemplo 10% = 0.10)",
]
df_empty = pd.DataFrame(columns=columns_order)

edited = st.data_editor(
    df_empty,
    num_rows="dynamic",
    use_container_width=True,
    hide_index=True,
    column_config={
        "ClasificaciÃ³n para el llenado de la declaraciÃ³n": st.column_config.SelectboxColumn(
            label="ClasificaciÃ³n para el llenado de la declaraciÃ³n",
            options=CLASIFICACIONES,
            required=True
        ),
        "Fecha de adquisiciÃ³n": st.column_config.DateColumn(label="Fecha de adquisiciÃ³n", format="YYYY-MM-DD"),
        "DescripciÃ³n": st.column_config.TextColumn(label="DescripciÃ³n"),
        "MOI": st.column_config.NumberColumn(label="MOI", format="%.2f", min_value=0.0, step=100.0),
        "Limite de la deducciÃ³n": st.column_config.NumberColumn(
            label="LÃ­mite de la deducciÃ³n (opcional)",
            format="%.2f", min_value=0.0, step=100.0,
            help="DÃ©jalo vacÃ­o para que lo determine la plantilla."
        ),
        "DepreciaciÃ³n acumulada": st.column_config.NumberColumn(label="DepreciaciÃ³n acumulada", format="%.2f", min_value=0.0, step=100.0),
        "Meses de Utilizacion": st.column_config.NumberColumn(label="Meses de UtilizaciÃ³n", min_value=0, max_value=12, step=1),
        "% de deducciÃ³n fiscal (capturar como factor ejemplo 10% = 0.10)": st.column_config.NumberColumn(
            label="% de deducciÃ³n fiscal (opcional)",
            format="%.2f %", min_value=0.0, step=0.01,
            help="Captura 10 para 10%. Si lo dejas vacÃ­o, la cÃ©dula usarÃ¡ el catÃ¡logo."
        ),
    }
)

# âââââââââââââââââââââââââââââ
# Generador: copiar plantilla y rellenar âDeduccion de Inversionesâ
# âââââââââââââââââââââââââââââ
def _fill_template_with_data(df_in: pd.DataFrame) -> bytes:
    if not TPL_CEDULA.exists():
        raise FileNotFoundError(f"No se encontrÃ³ la plantilla: {TPL_CEDULA}")
    wb = load_workbook(TPL_CEDULA, data_only=False)  # conserva fÃ³rmulas/formatos/hojas
    ws = wb["Deduccion de Inversiones"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    h2c = {h: i for i, h in enumerate(headers, start=1)}
    base_row = {c: ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)}

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    def put(r, c, v):
        cell = ws.cell(row=r, column=c)
        cell.value = v
        return cell  # mantiene formato del template

    n = len(df_in)
    start = 2
    for i in range(n):
        r = start + i
        ws.insert_rows(r)

        # Replicar fila 2 (ajustando referencias A2âAr, ...)
        for c in range(1, ws.max_column + 1):
            base = base_row[c]
            if isinstance(base, str) and base.startswith("="):
                nf = base
                for col_letter in list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")[:30]:
                    nf = nf.replace(f"{col_letter}2", f"{col_letter}{r}")
                put(r, c, nf)
            else:
                put(r, c, base)

        row = df_in.iloc[i]
        if "ClasificaciÃ³n para el llenado de la declaraciÃ³n" in h2c:
            put(r, h2c["ClasificaciÃ³n para el llenado de la declaraciÃ³n"], row.get("ClasificaciÃ³n para el llenado de la declaraciÃ³n"))
        if "Fecha de adquisiciÃ³n" in h2c:
            put(r, h2c["Fecha de adquisiciÃ³n"], row.get("Fecha de adquisiciÃ³n"))
        if "DescripciÃ³n" in h2c:
            put(r, h2c["DescripciÃ³n"], row.get("DescripciÃ³n"))

        if "MOI" in h2c:
            moi_val = row.get("MOI")
            lim_val = row.get("Limite de la deducciÃ³n")
            if (moi_val is None or moi_val == "") and (lim_val not in (None, "")):
                moi_val = lim_val
            put(r, h2c["MOI"], moi_val)

        if "Limite de la deducciÃ³n" in h2c and row.get("Limite de la deducciÃ³n") not in (None, ""):
            put(r, h2c["Limite de la deducciÃ³n"], row.get("Limite de la deducciÃ³n"))

        if "DepreciaciÃ³n acumulada" in h2c:
            put(r, h2c["DepreciaciÃ³n acumulada"], row.get("DepreciaciÃ³n acumulada"))

        if "Meses completos de utilizaciÃ³n" in h2c:
            put(r, h2c["Meses completos de utilizaciÃ³n"], row.get("Meses de Utilizacion"))

        h_col = "% de deducciÃ³n fiscal (capturar como factor ejemplo 10% = 0.10)"
        if h_col in h2c and row.get(h_col) not in (None, ""):
            put(r, h2c[h_col], _to_decimal_percent(row.get(h_col)))

    if n == 0:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=2, column=c).value = base_row[c]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# âââââââââââââââââââââââââââââ
# Un solo botÃ³n: genera y descarga
# âââââââââââââââââââââââââââââ
xls_bytes = _fill_template_with_data(edited)
file_name = f"Cedula_Deduccion_Anual_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

st.download_button(
    "ð¥ Descargar Excel (CÃ©dula de deducciÃ³n anual)",
    data=xls_bytes,
    file_name=file_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
