"""Generador para Pólizas Ig Cobranza con IVA integrado al hub principal."""

from __future__ import annotations

import io
from datetime import datetime
from urllib.parse import urlencode

import streamlit as st
from core.theme import apply_theme
from openpyxl import Workbook, load_workbook

from core.auth import ensure_session_from_token, auth_query_params
from core.custom_nav import handle_logout_request, render_brand_logout_nav


def _get_params() -> dict[str, str]:
    try:
        raw = st.query_params
    except Exception:
        raw = st.experimental_get_query_params()
    flattened: dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(value, list):
            value = value[-1] if value else None
        if value is None:
            continue
        flattened[key] = str(value)
    return flattened


def _handle_pending_navigation() -> None:
    params = _get_params()
    goto = params.pop("goto", None)
    if not goto:
        return
    try:
        st.query_params.clear()
        if params:
            st.query_params.update(params)
    except Exception:
        st.experimental_set_query_params(**params)
    try:
        st.switch_page(goto)
    except Exception:
        st.stop()
    st.stop()


def _back_href() -> str:
    params = {"goto": "pages/generador_polizas.py"}
    params.update(auth_query_params())
    query = urlencode(params, doseq=False)
    return f"?{query}"


st.set_page_config(
    page_title="Generador de pólizas Ig con IVA",
    layout="wide",
    initial_sidebar_state="collapsed",
)
apply_theme()
ensure_session_from_token()
handle_logout_request()
_handle_pending_navigation()

PAGE_CSS = """
<style>
#MainMenu, header[data-testid="stHeader"], footer, div[data-testid="stToolbar"], [data-testid="stSidebar"] {
  display:none !important;
}
.block-container { padding-top: 120px !important; max-width: 900px !important; }
body, [data-testid="stAppViewContainer"] { background:#f5f6fb !important; }
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

render_brand_logout_nav(
    "pages/generador_polizas.py",
    brand="Generador de pólizas",
    action_label="Atrás",
    action_href=_back_href(),
)

st.title("Generador de Polizas Ig Cobranza con IVA")


def parse_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            pass
    if "," in s and "." not in s:
        try:
            return float(s.replace(",", "."))
        except Exception:
            pass
    try:
        return float(s)
    except Exception:
        return v


def obtener_dia(valor):
    if isinstance(valor, datetime):
        return valor.day
    if isinstance(valor, (int, float)):
        return valor
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(str(valor), fmt).day
        except Exception:
            pass
    return valor


def escribir_bloque_base(ws, base, A, B, C, D, E, F, G, b3_val, codigo_b6, codigo_b7):
    ws[f"A{base}"] = "Ig"
    ws[f"C{base}"] = f'Cobranza Fact. {D if D is not None else ""}'
    ws[f"D{base}"] = obtener_dia(A)
    ws[f"B{base}"] = b3_val

    ws[f"B{base+1}"] = F
    ws[f"D{base+1}"] = G

    e_num = parse_number(E)
    ws[f"F{base+1}"] = e_num
    ws[f"F{base+1}"].number_format = "0.00"

    ws[f"G{base+2}"] = e_num
    ws[f"B{base+2}"] = B
    ws[f"D{base+2}"] = C

    ws[f"B{base+3}"] = codigo_b6
    ws[f"B{base+4}"] = codigo_b7
    ws[f"D{base+3}"] = "IVA por Trasladar"
    ws[f"D{base+4}"] = "IVA Trasladado"

    ws[f"F{base+3}"] = f"=F{base+1}/1.16*0.16"
    ws[f"G{base+4}"] = f"=F{base+1}/1.16*0.16"
    ws[f"F{base+3}"].number_format = "0.00"
    ws[f"G{base+4}"].number_format = "0.00"

    for cell in [f"C{base+1}", f"C{base+2}", f"C{base+3}", f"C{base+4}", f"G{base+1}", f"G{base+3}", f"F{base+2}", f"F{base+4}"]:
        ws[cell] = 0
    for cell in [f"E{base+1}", f"E{base+2}", f"E{base+3}", f"E{base+4}"]:
        ws[cell] = 1
    ws[f"B{base+5}"] = "FIN_PARTIDAS"


def escribir_bloque_433(ws, *args):
    return escribir_bloque_base(ws, *args, "2104-001-000", "2104-002-000")


def escribir_bloque_434(ws, *args):
    return escribir_bloque_base(ws, *args, "2104-001-0000", "2104-002-0000")


def escribir_bloque_436(ws, *args):
    return escribir_bloque_base(ws, *args, "2104-001-000000", "2104-002-000000")


def procesar_base(rows, b3_inicial, writer_func):
    dst_wb = Workbook()
    ws = dst_wb.active
    ws.title = "Hoja1"
    start_row, block_height, idx = 3, 6, 0

    for row in rows:
        A, B, C, D, E, F, G = (list(row) + [None] * 7)[:7]
        if all(v is None for v in [A, B, C, D, E, F, G]):
            continue
        base = start_row + idx * block_height
        b3_val = int(b3_inicial) + idx
        writer_func(ws, base, A, B, C, D, E, F, G, b3_val)
        idx += 1

    buf = io.BytesIO()
    dst_wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), idx


def get_b3_inicial():
    return int(st.session_state.get("b3_inicial", 1))


def get_uploaded_bytes():
    return st.session_state.get("uploaded_file_bytes", None)


def set_result(label, filename, data_bytes, blocks):
    st.session_state.update(
        {
            "result_label": label,
            "result_filename": filename,
            "result_bytes": data_bytes,
            "result_blocks": blocks,
        }
    )


st.number_input(
    "Número de Poliza Siguiente:",
    key="b3_inicial",
    min_value=0,
    value=1,
    step=1,
    format="%d",
)

uploaded_raw = st.file_uploader("Sube tu archivo Excel (xlsx)", type=["xlsx"], key="uploader_bottom")
if uploaded_raw is not None:
    st.session_state["uploaded_file_bytes"] = uploaded_raw.getvalue()

if get_uploaded_bytes():
    c1, c2, c3 = st.columns(3)
    if c1.button("Generar 4-3-3"):
        wb = load_workbook(io.BytesIO(get_uploaded_bytes()), data_only=False)
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        data, n = procesar_base(rows, get_b3_inicial(), escribir_bloque_433)
        set_result("4-3-3", "archivo_4-3-3.xlsx", data, n)
    if c2.button("Generar 4-3-4"):
        wb = load_workbook(io.BytesIO(get_uploaded_bytes()), data_only=False)
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        data, n = procesar_base(rows, get_b3_inicial(), escribir_bloque_434)
        set_result("4-3-4", "archivo_4-3-4.xlsx", data, n)
    if c3.button("Generar 4-3-6"):
        wb = load_workbook(io.BytesIO(get_uploaded_bytes()), data_only=False)
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        data, n = procesar_base(rows, get_b3_inicial(), escribir_bloque_436)
        set_result("4-3-6", "archivo_4-3-6.xlsx", data, n)
else:
    st.caption("👉 Adjunta tu archivo para habilitar los botones de generar.")

if "result_bytes" in st.session_state:
    st.success(f"✅ Procesado {st.session_state['result_label']}: {st.session_state['result_blocks']} bloque(s).")
    st.download_button(
        label=f"📥 Descargar {st.session_state['result_label']}",
        data=st.session_state["result_bytes"],
        file_name=st.session_state["result_filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
