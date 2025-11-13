"""Generador de pólizas Dr Egresos integrado al módulo principal."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Optional
from urllib.parse import urlencode

import streamlit as st
from core.theme import apply_theme
from openpyxl import Workbook, load_workbook

from core.auth import ensure_session_from_token, auth_query_params
from core.custom_nav import handle_logout_request, render_brand_logout_nav

try:
    from openpyxl.utils.datetime import from_excel as excel_serial_to_datetime  # type: ignore
except Exception:
    excel_serial_to_datetime = None


def _get_params() -> dict[str, str]:
    try:
        raw = st.query_params
    except Exception:
        raw = st.experimental_get_query_params()

    flattened: dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(value, list):
            value = value[-1] if value else None
        if value is None:
            continue
        flattened[key] = str(value)
    return flattened


def _handle_pending_navigation() -> None:
    params = _get_params()
    goto = params.pop("goto", None)
    if not goto:
        return
    try:
        st.query_params.clear()
        if params:
            st.query_params.update(params)
    except Exception:
        st.experimental_set_query_params(**params)
    try:
        st.switch_page(goto)
    except Exception:
        st.stop()
    st.stop()


def _back_href() -> str:
    params = {"goto": "pages/generador_polizas.py"}
    params.update(auth_query_params())
    query = urlencode(params, doseq=False)
    return f"?{query}"


st.set_page_config(page_title="Pólizas Dr Egresos", layout="wide", initial_sidebar_state="collapsed")
apply_theme()
ensure_session_from_token()
handle_logout_request()
_handle_pending_navigation()

PAGE_CSS = """
<style>
#MainMenu, header[data-testid="stHeader"], footer, div[data-testid="stToolbar"], [data-testid="stSidebar"] {
  display:none !important;
}
.block-container {
  padding-top: 120px !important;
  max-width: 900px !important;
}
body, [data-testid="stAppViewContainer"] {
  background:#f5f6fb !important;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

render_brand_logout_nav(
    "pages/generador_polizas.py",
    brand="Generador de pólizas",
    action_label="Atrás",
    action_href=_back_href(),
)


def extract_day(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.day
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(v.strip(), fmt).day
            except Exception:
                pass
        return None
    if isinstance(v, (int, float)) and excel_serial_to_datetime is not None:
        try:
            dt = excel_serial_to_datetime(v)
            return dt.day if isinstance(dt, datetime) else None
        except Exception:
            return None
    return None


def is_diff_zero(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        try:
            return float(v) != 0.0
        except Exception:
            return True
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return False
        s_norm = s.replace(",", ".")
        try:
            return float(s_norm) != 0.0
        except Exception:
            return s != "0"
    return True


def has_signal_for_row(D, E, G) -> bool:
    return any(x not in (None, "") for x in (D, E, G))


def write_row(ws, r, B=None, C=None, D=None, E=None, F=None, G=None):
    if B is not None:
        ws[f"B{r}"] = B
    if C is not None:
        ws[f"C{r}"] = C
    if D is not None:
        ws[f"D{r}"] = D
    if E is not None:
        ws[f"E{r}"] = E
    if F is not None:
        ws[f"F{r}"] = F
    if G is not None:
        ws[f"G{r}"] = G


def build_output_workbook_batch(ws_src, seq_start: int) -> Workbook:
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Hoja1"

    last = ws_src.max_row
    cursor = 3
    seq = int(seq_start)

    for r in range(2, last + 1):
        A = ws_src[f"A{r}"].value
        B = ws_src[f"B{r}"].value
        C = ws_src[f"C{r}"].value
        D = ws_src[f"D{r}"].value
        E = ws_src[f"E{r}"].value
        F = ws_src[f"F{r}"].value
        G = ws_src[f"G{r}"].value
        H = ws_src[f"H{r}"].value
        I = ws_src[f"I{r}"].value
        J = ws_src[f"J{r}"].value
        K = ws_src[f"K{r}"].value
        L = ws_src[f"L{r}"].value
        M = ws_src[f"M{r}"].value
        N = ws_src[f"N{r}"].value

        if not has_signal_for_row(D, E, G) and not any(is_diff_zero(x) for x in (B, I, J, K, L, M)):
            continue

        ws_out[f"A{cursor}"] = "Dr"
        ws_out[f"B{cursor}"] = seq
        ws_out[f"C{cursor}"] = f"Provisión Egresos Fact. {'' if F is None else F}"
        day = extract_day(A)
        ws_out[f"D{cursor}"] = day if day is not None else "DIA_PENDIENTE"

        row1 = cursor + 1
        if has_signal_for_row(D, E, G):
            write_row(
                ws_out,
                row1,
                B=D,
                C=0,
                D=E if E is not None else "",
                E=1,
                F=G if G is not None else "",
                G=0,
            )
            next_row = row1 + 1
        else:
            next_row = row1

        write_row(
            ws_out,
            next_row,
            B="1104-0001-000000",
            C=0,
            D="IVA por Acreditar",
            E=1,
            F=H if H is not None else "",
            G=0,
        )
        next_row += 1

        if is_diff_zero(I):
            write_row(ws_out, next_row, B="2103-0009-000000", C=0, D="IVA Ret  x Serv Profesionales", E=1, F=0, G=I)
            next_row += 1
        if is_diff_zero(J):
            write_row(ws_out, next_row, B="2103-0012-000000", C=0, D="Retención de IVA 4%", E=1, F=0, G=J)
            next_row += 1
        if is_diff_zero(K):
            write_row(ws_out, next_row, B="2103-0011-000000", C=0, D="IVA Ret  x Arrendamiento", E=1, F=0, G=K)
            next_row += 1
        if is_diff_zero(L):
            write_row(ws_out, next_row, B="2103-0008-000000", C=0, D="ISR Ret  x Serv Profesionales", E=1, F=0, G=L)
            next_row += 1
        if is_diff_zero(M):
            write_row(ws_out, next_row, B="2103-0010-000000", C=0, D="ISR Ret x Arrendamiento", E=1, F=0, G=M)
            next_row += 1
        if is_diff_zero(B):
            write_row(ws_out, next_row, B=B, C=0, D=C if C is not None else "", E=1, F=0, G=N if N is not None else "")
            next_row += 1

        ws_out[f"B{next_row}"] = "FIN_PARTIDAS"
        cursor = next_row + 1
        seq += 1

    return wb_out


def process_source_excel(file_bytes: bytes, seq_start: int):
    wb_src = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws_src = wb_src.active
    wb_out = build_output_workbook_batch(ws_src, seq_start)
    out = io.BytesIO()
    wb_out.save(out)
    return out.getvalue()


st.title("📄 Pólizas Dr Egresos")

seq_start = st.number_input(
    "Valor inicial para B (encabezado de cada bloque)",
    min_value=0,
    value=1,
    step=1,
)

uploaded = st.file_uploader("Selecciona tu archivo .xlsx", type=["xlsx"])

if uploaded is not None:
    try:
        out_bytes = process_source_excel(uploaded.read(), seq_start=int(seq_start))
        st.success("✅ Archivo generado correctamente.")
        st.download_button(
            label="⬇️ Descargar Excel generado",
            data=out_bytes,
            file_name="poliza_dr_egresos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        st.error("Ocurrió un error al procesar el archivo.")
        st.exception(exc)
else:
    st.info("Carga un archivo .xlsx para comenzar.")
