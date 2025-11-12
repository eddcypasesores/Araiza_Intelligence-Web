"""Generador para pólizas Ig Cobranza sin IVA."""

from __future__ import annotations

import io
from datetime import datetime
from urllib.parse import urlencode

import streamlit as st
from openpyxl import Workbook, load_workbook

from core.auth import ensure_session_from_token, auth_query_params
from core.custom_nav import handle_logout_request, render_brand_logout_nav


def _get_params() -> dict[str, str]:
    try:
        raw = st.query_params
    except Exception:
        raw = st.experimental_get_query_params()
    flattened: dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(value, list):
            value = value[-1] if value else None
        if value is None:
            continue
        flattened[key] = str(value)
    return flattened


def _handle_pending_navigation() -> None:
    params = _get_params()
    goto = params.pop("goto", None)
    if not goto:
        return
    try:
        st.query_params.clear()
        if params:
            st.query_params.update(params)
    except Exception:
        st.experimental_set_query_params(**params)
    try:
        st.switch_page(goto)
    except Exception:
        st.stop()
    st.stop()


def _back_href() -> str:
    params = {"goto": "pages/generador_polizas.py"}
    params.update(auth_query_params())
    query = urlencode(params, doseq=False)
    return f"?{query}"


st.set_page_config(
    page_title="Generador de pólizas Ig sin IVA",
    layout="wide",
    initial_sidebar_state="collapsed",
)
ensure_session_from_token()
handle_logout_request()
_handle_pending_navigation()

PAGE_CSS = """
<style>
#MainMenu, header[data-testid="stHeader"], footer, div[data-testid="stToolbar"], [data-testid="stSidebar"] {
  display:none !important;
}
.block-container { padding-top: 120px !important; max-width: 900px !important; }
body, [data-testid="stAppViewContainer"] { background:#f5f6fb !important; }
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

render_brand_logout_nav(
    "pages/generador_polizas.py",
    brand="Generador de pólizas",
    action_label="Atrás",
    action_href=_back_href(),
)

st.title("Generador de Polizas Ig Cobranza sin IVA")


def parse_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            pass
    if "," in s and "." not in s:
        try:
            return float(s.replace(",", "."))
        except Exception:
            pass
    try:
        return float(s)
    except Exception:
        return v


def obtener_dia(valor):
    if isinstance(valor, datetime):
        return valor.day
    if isinstance(valor, (int, float)):
        return valor
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(str(valor), fmt).day
        except Exception:
            pass
    return valor


def escribir_bloque_sin_iva(ws, base, A, B, C, D, E, F, G, b3_val):
    ws[f"A{base}"] = "Ig"
    ws[f"C{base}"] = f'Cobranza Fact. {D if D is not None else ""}'
    ws[f"D{base}"] = obtener_dia(A)
    ws[f"B{base}"] = b3_val

    ws[f"B{base+1}"] = F
    ws[f"D{base+1}"] = G
    e_num = parse_number(E)
    ws[f"F{base+1}"] = e_num
    ws[f"F{base+1}"].number_format = "0.00"

    ws[f"G{base+2}"] = e_num
    ws[f"B{base+2}"] = B
    ws[f"D{base+2}"] = C

    for cell in [f"C{base+1}", f"C{base+2}", f"G{base+1}", f"F{base+2}"]:
        ws[cell] = 0
    for cell in [f"E{base+1}", f"E{base+2}"]:
        ws[cell] = 1

    ws[f"B{base+3}"] = "FIN_PARTIDAS"


def procesar(rows, b3_inicial):
    dst_wb = Workbook()
    ws = dst_wb.active
    ws.title = "Hoja1"

    start_row = 3
    block_height = 4
    idx = 0
    for row in rows:
        A, B, C, D, E, F, G = (list(row) + [None] * 7)[:7]
        if all(v is None for v in [A, B, C, D, E, F, G]):
            continue
        base = start_row + idx * block_height
        b3_val = int(b3_inicial) + idx
        escribir_bloque_sin_iva(ws, base, A, B, C, D, E, F, G, b3_val)
        idx += 1

    buf = io.BytesIO()
    dst_wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), idx


def get_b3_inicial():
    return int(st.session_state.get("b3_inicial", 1))


def get_uploaded_bytes():
    return st.session_state.get("uploaded_file_bytes", None)


st.number_input(
    "Número de póliza siguiente",
    key="b3_inicial",
    min_value=0,
    value=1,
    step=1,
    format="%d",
)

uploaded = st.file_uploader("Sube tu archivo Excel (xlsx)", type=["xlsx"], key="uploader_bottom")
if uploaded is not None:
    st.session_state["uploaded_file_bytes"] = uploaded.getvalue()

if get_uploaded_bytes():
    wb = load_workbook(io.BytesIO(get_uploaded_bytes()), data_only=False)
    rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    data_bytes, _ = procesar(rows, get_b3_inicial())

    st.download_button(
        label="📥 Descargar Excel",
        data=data_bytes,
        file_name="polizas_ig_sin_iva.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Sube un archivo .xlsx para comenzar.")
