"""Generador para Pólizas Eg Pago a Proveedores."""

from __future__ import annotations

from datetime import datetime, date
from io import BytesIO
from urllib.parse import urlencode

import streamlit as st
from core.theme import apply_theme
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel as from_excel_serial

from core.auth import ensure_session_from_token, auth_query_params
from core.custom_nav import handle_logout_request, render_brand_logout_nav


def _get_params() -> dict[str, str]:
    try:
        raw = st.query_params
    except Exception:
        raw = st.experimental_get_query_params()

    flattened: dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(value, list):
            value = value[-1] if value else None
        if value is None:
            continue
        flattened[key] = str(value)
    return flattened


def _handle_pending_navigation() -> None:
    params = _get_params()
    goto = params.pop("goto", None)
    if not goto:
        return
    try:
        st.query_params.clear()
        if params:
            st.query_params.update(params)
    except Exception:
        st.experimental_set_query_params(**params)
    try:
        st.switch_page(goto)
    except Exception:
        st.stop()
    st.stop()


def _back_href() -> str:
    params = {"goto": "pages/generador_polizas.py"}
    params.update(auth_query_params())
    query = urlencode(params, doseq=False)
    return f"?{query}"


st.set_page_config(
    page_title="Polizas Eg",
    layout="wide",
    initial_sidebar_state="collapsed",
)
apply_theme()
ensure_session_from_token()
handle_logout_request()
_handle_pending_navigation()

PAGE_CSS = """
<style>
#MainMenu, header[data-testid="stHeader"], footer, div[data-testid="stToolbar"], [data-testid="stSidebar"] {
  display:none !important;
}
.block-container {
  padding-top: 120px !important;
  max-width: 900px !important;
}
body, [data-testid="stAppViewContainer"] {
  background:#f5f6fb !important;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

render_brand_logout_nav(
    "pages/generador_polizas.py",
    brand="Generador de pólizas",
    action_label="Atrás",
    action_href=_back_href(),
)


st.title("Generador de Polizas Eg Pago a Proveedores")

consecutivo_inicial = st.number_input(
    "Consecutivo inicial para B3 (se incrementa +1 por bloque)",
    min_value=0,
    value=1,
    step=1,
)

uploaded = st.file_uploader("📎 Sube tu Excel fuente (.xlsx)", type=["xlsx"])


def extract_day(value):
    if isinstance(value, (datetime, date)):
        return value.day
    if isinstance(value, (int, float)):
        try:
            return from_excel_serial(value).day
        except Exception:
            return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d", "%d/%m", "%d-%m"):
            try:
                return datetime.strptime(s, fmt).day
            except Exception:
                pass
        return value
    return value


def is_nonzero(val) -> bool:
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return abs(float(val)) > 1e-9
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return False
        try:
            return abs(float(s.replace(",", ""))) > 1e-9
        except Exception:
            return s != "0"
    return True


def not_literal_zero(val) -> bool:
    if val is None:
        return False
    return str(val).strip() != "0"


def to_number_if_possible(v):
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except Exception:
            return v
    return v


def write_row(ws, row, B=None, C=None, D=None, E=None, F=None, G=None):
    if B is not None:
        ws[f"B{row}"] = B
    if C is not None:
        ws[f"C{row}"] = C
    if D is not None:
        ws[f"D{row}"] = D
    if E is not None:
        ws[f"E{row}"] = E
    if F is not None:
        ws[f"F{row}"] = F
    if G is not None:
        ws[f"G{row}"] = G


def generar_excel_salida(wb_src_bytes: bytes, consecutivo_b3_inicial: int) -> bytes:
    wb_src = load_workbook(BytesIO(wb_src_bytes), data_only=True)
    ws_src = wb_src.active

    wb_dst = Workbook()
    ws_dst = wb_dst.active

    out_row = 3
    consecutivo = int(consecutivo_b3_inicial)

    max_row = ws_src.max_row if ws_src.max_row else 2
    for r in range(2, max_row + 1):
        A = ws_src[f"A{r}"].value
        B = ws_src[f"B{r}"].value
        C = ws_src[f"C{r}"].value
        D = ws_src[f"D{r}"].value
        E = ws_src[f"E{r}"].value
        F = ws_src[f"F{r}"].value
        G = ws_src[f"G{r}"].value
        H = ws_src[f"H{r}"].value
        N = ws_src[f"N{r}"].value
        O = ws_src[f"O{r}"].value
        P = ws_src[f"P{r}"].value
        Q = ws_src[f"Q{r}"].value
        R = ws_src[f"R{r}"].value

        if all(v in (None, "") for v in (A, B, C, D, E, F, G, H, N, O, P, Q, R)):
            continue

        ws_dst[f"A{out_row}"] = "Eg"
        ws_dst[f"B{out_row}"] = consecutivo
        ws_dst[f"C{out_row}"] = f"Pago Fact. {'' if F is None else str(F)}"
        ws_dst[f"D{out_row}"] = extract_day(A)

        write_row(
            ws_dst,
            out_row + 1,
            B="" if B is None else str(B),
            C=0,
            D="" if C is None else str(C),
            E=1,
            F=N,
            G=0,
        )

        write_row(
            ws_dst,
            out_row + 2,
            B="" if D is None else str(D),
            C=0,
            D="" if E is None else str(E),
            E=1,
            F=0,
            G=N,
        )

        cursor = out_row + 3

        if is_nonzero(H):
            h_val = to_number_if_possible(H)
            write_row(
                ws_dst,
                cursor,
                B="1106-0001-000000",
                C=0,
                D="IVA Acrediatable",
                E=1,
                F=h_val,
                G=0,
            )
            cursor += 1
            write_row(
                ws_dst,
                cursor,
                B="1106-0002-000000",
                C=0,
                D="IVA por Acreditar",
                E=1,
                F=0,
                G=h_val,
            )
            cursor += 1

        if is_nonzero(O):
            g_val = to_number_if_possible(G)
            write_row(
                ws_dst,
                cursor,
                B=O,
                C=0,
                D=P,
                E=1,
                F=g_val,
                G=0,
            )
            cursor += 1
            write_row(
                ws_dst,
                cursor,
                B=Q,
                C=0,
                D=R,
                E=1,
                F=0,
                G=g_val,
            )
            cursor += 1

        if not_literal_zero(B):
            write_row(ws_dst, cursor, B="FIN_PARTIDAS")
            cursor += 1

        out_row = cursor
        consecutivo += 1

    bio = BytesIO()
    wb_dst.save(bio)
    bio.seek(0)
    return bio.read()


if uploaded:
    try:
        out_bytes = generar_excel_salida(uploaded.read(), consecutivo_inicial)
        st.success("✅ ¡Excel generado con éxito!")
        st.download_button(
            label="⬇️ Descargar Excel generado",
            data=out_bytes,
            file_name="Poliza_Generada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"❌ Ocurrió un error al procesar el archivo: {e}")
else:
    st.info("Sube un archivo .xlsx para comenzar.")
