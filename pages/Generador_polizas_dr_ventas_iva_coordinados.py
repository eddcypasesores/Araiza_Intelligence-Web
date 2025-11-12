"""Generador para Pólizas Dr Ventas con IVA Coordinados."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Optional
from urllib.parse import urlencode

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from core.auth import ensure_session_from_token, auth_query_params
from core.custom_nav import handle_logout_request, render_brand_logout_nav


def _get_params() -> dict[str, str]:
    try:
        raw = st.query_params
    except Exception:
        raw = st.experimental_get_query_params()

    flattened: dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(value, list):
            value = value[-1] if value else None
        if value is None:
            continue
        flattened[key] = str(value)
    return flattened


def _handle_pending_navigation() -> None:
    params = _get_params()
    goto = params.pop("goto", None)
    if not goto:
        return
    try:
        st.query_params.clear()
        if params:
            st.query_params.update(params)
    except Exception:
        st.experimental_set_query_params(**params)
    try:
        st.switch_page(goto)
    except Exception:
        st.stop()
    st.stop()


def _back_href() -> str:
    params = {"goto": "pages/generador_polizas.py"}
    params.update(auth_query_params())
    query = urlencode(params, doseq=False)
    return f"?{query}"


st.set_page_config(
    page_title="Pólizas Dr Ventas con IVA Coordinados",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed",
)
ensure_session_from_token()
handle_logout_request()
_handle_pending_navigation()

PAGE_CSS = """
<style>
#MainMenu, header[data-testid="stHeader"], footer, div[data-testid="stToolbar"], [data-testid="stSidebar"] {
  display:none !important;
}
.block-container {
  padding-top: 120px !important;
  max-width: 900px !important;
}
body, [data-testid="stAppViewContainer"] {
  background:#f5f6fb !important;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

render_brand_logout_nav(
    "pages/generador_polizas.py",
    brand="Generador de pólizas",
    action_label="Atrás",
    action_href=_back_href(),
)


def extract_day(value):
    if isinstance(value, datetime):
        return value.day
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
            return dt.day
        except Exception:
            try:
                return int(value)
            except Exception:
                return value
    if isinstance(value, str):
        fmts = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"]
        for f in fmts:
            try:
                return datetime.strptime(value.strip(), f).day
            except Exception:
                pass
        return value
    return value


def fila_vacia(valores):
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in valores)


def parse_number(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.replace("$", "").replace(",", "").replace(" ", "")
        try:
            return float(s)
        except Exception:
            return None
    return None


def number_format_from_decimals(decimals: int) -> str:
    if decimals <= 0:
        return "0"
    return "0." + ("0" * decimals)


def write_number(cell, value, decimals: int):
    cell.value = value
    cell.number_format = number_format_from_decimals(decimals)


def is_zero_like(val) -> bool:
    if val is None:
        return True
    if isinstance(val, (int, float)):
        return abs(val) < 1e-9
    return False


def prune_zero_rows(ws_out, preserve_prefix: str = "1103-"):
    row = 1
    while row <= ws_out.max_row:
        a = ws_out.cell(row=row, column=1).value
        b = ws_out.cell(row=row, column=2).value
        f = ws_out.cell(row=row, column=6).value
        g = ws_out.cell(row=row, column=7).value

        if a is None:
            if not (isinstance(b, str) and b.startswith(preserve_prefix)):
                if is_zero_like(f) and is_zero_like(g):
                    ws_out.delete_rows(row, 1)
                    continue
        row += 1


def write_block(
    ws_out,
    base_row,
    src_A,
    src_B,
    src_C,
    src_D,
    src_E,
    src_F,
    src_G,
    src_H,
    consecutivo: int,
    cod_ingresos: str,
    cod_iva_tras: str,
    cod_iva_ret: str,
    desc_ingresos: str,
    desc_iva_tras: str,
    desc_iva_ret: str,
    incluir_fila7: bool,
    decimales: int,
) -> int:
    r = base_row

    ws_out.cell(row=r, column=1, value="Dr.")
    ws_out.cell(row=r, column=2, value=int(consecutivo))
    ws_out.cell(row=r, column=3, value=f"Provisión Fact. {'' if src_D is None else src_D}")
    ws_out.cell(row=r, column=4, value=extract_day(src_A))

    ws_out.cell(row=r + 1, column=2, value=src_B)
    ws_out.cell(row=r + 1, column=3, value=0)
    ws_out.cell(row=r + 1, column=4, value=src_C)
    ws_out.cell(row=r + 1, column=5, value=1)
    val_h = parse_number(src_H) or 0.0
    write_number(ws_out.cell(row=r + 1, column=6), val_h, decimales)
    write_number(ws_out.cell(row=r + 1, column=7), 0.0, decimales)

    ws_out.cell(row=r + 2, column=2, value=cod_ingresos)
    ws_out.cell(row=r + 2, column=3, value=0)
    ws_out.cell(row=r + 2, column=4, value=desc_ingresos)
    ws_out.cell(row=r + 2, column=5, value=1)
    write_number(ws_out.cell(row=r + 2, column=6), 0.0, decimales)
    val_e = parse_number(src_E) or 0.0
    write_number(ws_out.cell(row=r + 2, column=7), val_e, decimales)

    ws_out.cell(row=r + 3, column=2, value=cod_iva_tras)
    ws_out.cell(row=r + 3, column=3, value=0)
    ws_out.cell(row=r + 3, column=4, value=desc_iva_tras)
    ws_out.cell(row=r + 3, column=5, value=1)
    write_number(ws_out.cell(row=r + 3, column=6), 0.0, decimales)
    val_f = parse_number(src_F) or 0.0
    write_number(ws_out.cell(row=r + 3, column=7), val_f, decimales)

    if incluir_fila7:
        ws_out.cell(row=r + 4, column=2, value=cod_iva_ret)
        ws_out.cell(row=r + 4, column=3, value=0)
        ws_out.cell(row=r + 4, column=4, value=desc_iva_ret)
        ws_out.cell(row=r + 4, column=5, value=1)
        val_g = parse_number(src_G) or 0.0
        write_number(ws_out.cell(row=r + 4, column=6), val_g, decimales)
        write_number(ws_out.cell(row=r + 4, column=7), 0.0, decimales)
        ws_out.cell(row=r + 5, column=2, value="FIN_PARTIDAS")
        return 6
    else:
        ws_out.cell(row=r + 4, column=2, value="FIN_PARTIDAS")
        return 5


def build_stacked_output(
    src_bytes: bytes,
    consecutivo_inicial: int,
    cod_ingresos="4101-001-000000",
    cod_iva_tras="2104-001-000000",
    cod_iva_ret="1106-001-000000",
    desc_ingresos="Ingresos por coordinados 16%",
    desc_iva_tras="IVA trasladado no cobrado",
    desc_iva_ret="IVA Retenido",
    incluir_fila7=True,
    eliminar_filas_cero=False,
    decimales=2,
    sheet_title="Hoja1",
) -> bytes:
    wb_src = load_workbook(filename=BytesIO(src_bytes), data_only=True)
    ws_src = wb_src.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = sheet_title

    current = 3
    max_row = ws_src.max_row
    generadas = 0
    consecutivo = int(consecutivo_inicial)

    for r in range(2, max_row + 1):
        A = ws_src.cell(row=r, column=1).value
        B = ws_src.cell(row=r, column=2).value
        C = ws_src.cell(row=r, column=3).value
        D = ws_src.cell(row=r, column=4).value
        E = ws_src.cell(row=r, column=5).value
        F = ws_src.cell(row=r, column=6).value
        G = ws_src.cell(row=r, column=7).value
        H = ws_src.cell(row=r, column=8).value

        if fila_vacia((A, B, C, D, E)):
            continue

        wrote = write_block(
            ws_out,
            base_row=current,
            src_A=A,
            src_B=B,
            src_C=C,
            src_D=D,
            src_E=E,
            src_F=F,
            src_G=G,
            src_H=H,
            consecutivo=consecutivo,
            cod_ingresos=cod_ingresos,
            cod_iva_tras=cod_iva_tras,
            cod_iva_ret=cod_iva_ret,
            desc_ingresos=desc_ingresos,
            desc_iva_tras=desc_iva_tras,
            desc_iva_ret=desc_iva_ret,
            incluir_fila7=incluir_fila7,
            decimales=decimales,
        )
        current += wrote
        consecutivo += 1
        generadas += 1

    if generadas == 0:
        raise ValueError("No se encontraron filas con datos desde la fila 2 en las columnas A–E.")

    if eliminar_filas_cero:
        prune_zero_rows(ws_out, preserve_prefix="1103-")

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.getvalue()


def render_tab(
    tab_label,
    cod_ingresos_default,
    cod_iva_tras_default,
    cod_iva_ret_default,
):
    st.subheader(f"Proceso {tab_label}")
    uploaded = st.file_uploader(
        "Selecciona un archivo Excel (.xlsx)",
        type=["xlsx"],
        key=f"uploader_{tab_label}",
    )
    consec_ini = st.number_input(
        "Consecutivo inicial (B3 del primer proceso)",
        min_value=0,
        value=1,
        step=1,
        key=f"consec_{tab_label}",
    )

    with st.expander("⚙️ Opciones"):
        incluir_fila7 = st.checkbox(
            "Incluir Fila 7 (IVA Retenido)",
            value=True,
            key=f"fila7_{tab_label}",
        )
        eliminar_filas_cero = st.checkbox(
            "Eliminar filas donde F=0 y G=0 (excepto B inicia con 1103-)",
            value=False,
            key=f"elim0_{tab_label}",
        )
        decimales = st.number_input(
            "Decimales para importes",
            min_value=0,
            max_value=6,
            value=2,
            step=1,
            key=f"decs_{tab_label}",
        )
        sheet_title = st.text_input(
            "Título de la hoja de salida",
            value="Hoja1",
            key=f"title_{tab_label}",
        )

        st.markdown("**Códigos (B5/B6/B7)**")
        cod_ingresos = st.text_input(
            "Código ingresos (B5)",
            value=cod_ingresos_default,
            key=f"cod_ing_{tab_label}",
        )
        cod_iva_tras = st.text_input(
            "Código IVA trasladado (B6)",
            value=cod_iva_tras_default,
            key=f"cod_tras_{tab_label}",
        )
        cod_iva_ret = st.text_input(
            "Código IVA retenido (B7)",
            value=cod_iva_ret_default,
            key=f"cod_ret_{tab_label}",
        )

        st.markdown("**Descripciones (D5/D6/D7)**")
        desc_ingresos = st.text_input(
            "Descripción ingresos (D5)",
            value="Ingresos por coordinados 16%",
            key=f"desc_ing_{tab_label}",
        )
        desc_iva_tras = st.text_input(
            "Descripción IVA trasladado (D6)",
            value="IVA trasladado no cobrado",
            key=f"desc_tras_{tab_label}",
        )
        desc_iva_ret = st.text_input(
            "Descripción IVA retenido (D7)",
            value="IVA Retenido",
            key=f"desc_ret_{tab_label}",
        )

    if uploaded:
        if st.button("Generar archivo apilado", key=f"btn_{tab_label}"):
            try:
                data = build_stacked_output(
                    uploaded.getvalue(),
                    consecutivo_inicial=consec_ini,
                    cod_ingresos=cod_ingresos,
                    cod_iva_tras=cod_iva_tras,
                    cod_iva_ret=cod_iva_ret,
                    desc_ingresos=desc_ingresos,
                    desc_iva_tras=desc_iva_tras,
                    desc_iva_ret=desc_iva_ret,
                    incluir_fila7=incluir_fila7,
                    eliminar_filas_cero=eliminar_filas_cero,
                    decimales=decimales,
                    sheet_title=sheet_title,
                )
                st.success(f"Archivo generado con todos los procesos apilados ({tab_label}).")
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=data,
                    file_name=f"Salidas_Apiladas_{tab_label}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{tab_label}",
                )
            except Exception as e:
                st.error(f"Ocurrió un error: {e}")
        with st.expander("Vista rápida (fila 2)"):
            try:
                wb_tmp = load_workbook(filename=BytesIO(uploaded.getvalue()), data_only=True)
                ws_tmp = wb_tmp.active
                st.write(
                    {
                        "A2 (para día)": ws_tmp["A2"].value,
                        "B2 → B4": ws_tmp["B2"].value,
                        "C2 → D4": ws_tmp["C2"].value,
                        "D2 → C3": ws_tmp["D2"].value,
                        "E2 → G5": ws_tmp["E2"].value,
                        "F2 → G6": ws_tmp["F2"].value,
                        "G2 → F7": ws_tmp["G2"].value,
                        "H2 → F4": ws_tmp["H2"].value,
                        "Última fila detectada": ws_tmp.max_row,
                    }
                )
            except Exception as e:
                st.warning(f"No se pudieron mostrar detalles: {e}")
    else:
        st.info("Esperando un archivo .xlsx…")


st.title("Dr Ventas con IVA Coordinados")

tab_433, tab_434, tab_436, tab_446 = st.tabs(["4-3-3", "4-3-4", "4-3-6", "4-4-6"])

with tab_433:
    render_tab(
        "4-3-3",
        cod_ingresos_default="4101-001-000",
        cod_iva_tras_default="2104-001-000",
        cod_iva_ret_default="1106-001-000",
    )

with tab_434:
    render_tab(
        "4-3-4",
        cod_ingresos_default="4101-001-0000",
        cod_iva_tras_default="2104-001-0000",
        cod_iva_ret_default="1106-001-0000",
    )

with tab_436:
    render_tab(
        "4-3-6",
        cod_ingresos_default="4101-001-000000",
        cod_iva_tras_default="2104-001-000000",
        cod_iva_ret_default="1106-001-000000",
    )

with tab_446:
    render_tab(
        "4-4-6",
        cod_ingresos_default="4101-0001-000000",
        cod_iva_tras_default="2104-0001-000000",
        cod_iva_ret_default="1106-0001-000000",
    )
