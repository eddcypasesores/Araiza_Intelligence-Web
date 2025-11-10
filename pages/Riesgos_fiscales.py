from __future__ import annotations

import io
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile, ZipFile

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.auth import ensure_session_from_token, persist_login
from core.custom_nav import handle_logout_request, render_brand_logout_nav
from core.db import authenticate_portal_user, ensure_schema, get_conn
from core.flash import consume_flash
from core.login_ui import render_login_header, render_token_reset_section
from core.streamlit_compat import set_query_params

# ===================== Config y controles de sesión =====================
st.set_page_config(page_title="Riesgos Fiscales", layout="centered")

ensure_session_from_token()
handle_logout_request()

MODULE_PERMISSION = "riesgos"
MODULE_TITLE = "Riesgos fiscales"
RESET_SLUG = "riesgos"


def _has_permission() -> bool:
    permisos = set(st.session_state.get("permisos") or [])
    return MODULE_PERMISSION in permisos or "admin" in permisos


def _render_login() -> None:
    consume_flash()
    render_login_header("Iniciar sesión", subtitle=MODULE_TITLE)
    st.caption("Valida tus credenciales para acceder al monitoreo especializado de riesgos fiscales.")

    with st.form("riesgos_login", clear_on_submit=False):
        username = st.text_input("RFC", placeholder="ej. ABCD800101XXX")
        password = st.text_input("Contraseña", type="password", placeholder="********")
        col_login, col_cancel = st.columns(2)
        submitted = col_login.form_submit_button("Iniciar sesión", use_container_width=True)
        cancelled = col_cancel.form_submit_button("Cancelar", use_container_width=True)

    if cancelled:
        st.switch_page("pages/0_Inicio.py")
        st.stop()

    handled_reset = render_token_reset_section(RESET_SLUG)
    if handled_reset:
        st.stop()

    if not submitted:
        st.stop()

    username = (username or "").strip().upper()
    password = password or ""

    conn = get_conn()
    ensure_schema(conn)
    try:
        record = authenticate_portal_user(conn, username, password)
    except Exception as exc:
        st.error("No fue posible validar las credenciales. Inténtalo de nuevo.")
        st.caption(f"Detalle técnico: {exc}")
        st.stop()
    finally:
        conn.close()

    if not record:
        st.error("RFC o contraseña incorrectos.")
        st.stop()

    permisos = set(record.get("permisos") or [])
    if MODULE_PERMISSION not in permisos and "admin" not in permisos:
        st.error("Tu cuenta no tiene permiso para acceder a Riesgos fiscales.")
        st.stop()

    token = persist_login(
        record["rfc"],
        record["permisos"],
        must_change_password=record.get("must_change_password", False),
        user_id=record.get("id"),
    )

    try:
        params = {k: v for k, v in st.query_params.items() if k != "auth"}
        params["auth"] = token
        set_query_params(params)
    except Exception:
        pass

    st.rerun()


# ===================== Estilos mínimos =====================
st.markdown(
    """
    <style>
      html, body, .stApp, .block-container { background:#ffffff !important; color:#111 !important; }
      #MainMenu, header, footer, div[data-testid=\"stToolbar\"] { display:none !important; }
      .block-container { padding-top: 1rem; }
      *, *::before, *::after { color:#111 !important; opacity:1 !important; }
      h1, h2, h3, h4, h5, h6, label { color:#111 !important; }
      div[data-testid=\"stTextInput\"] input{
        padding:.5rem .9rem; font-size:.95rem; color:#111 !important;
        background:#fff !important; border:1px solid #d0d7de !important; border-radius:.5rem !important;
      }
      input::placeholder { color:#9aa0a6 !important; opacity:1 !important; }
      .small-error { color:#b91c1c; font-size:.85rem; margin-top:.25rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

if not (st.session_state.get("usuario") and _has_permission()):
    _render_login()
    st.stop()

render_brand_logout_nav(MODULE_TITLE)
st.title("RIESGOS FISCALES")

# ===================== Inputs arriba =====================
st.markdown(
    """
    <style>
      .riesgos-card {
        border:1px solid #e2e8f0;
        border-radius:24px;
        padding:20px clamp(16px,2vw,28px);
        background:#fff;
        box-shadow:0 18px 40px rgba(15,23,42,.08);
        margin-bottom:1.2rem;
      }
      .riesgos-card h3 {
        margin-top:0;
        font-size:1rem;
        font-weight:800;
        letter-spacing:.08em;
        color:#0f172a;
      }
      .uploader-label {
        font-size:.9rem;
        font-weight:600;
        margin-bottom:.25rem;
      }
      .uploader-centered {
        text-align:center;
        display:block;
      }
      div[data-testid="stFileUploader"]>label { display:none; }
      div[data-testid="stFileUploader"] div[data-testid="stFileUploaderDropzone"] {
        border-radius:14px;
        border:1px dashed #94a3b8;
        background:#f8fafc;
        min-height:140px;
        width:100%;
        padding:1.1rem 1.2rem;
      }
      div[data-testid="stFileUploader"] div[data-testid="stFileUploaderDropzone"] > div {
        justify-content:center;
        width:100%;
      }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown("<div class='riesgos-card'>", unsafe_allow_html=True)
form_cols = st.columns([1.4, 1.2, 0.6])
with form_cols[0]:
    regimen_fiscal_correcto = st.text_input("Régimen fiscal correcto", value="", placeholder="Ej. 601")
with form_cols[1]:
    cp_val = st.text_input("CP", value="", placeholder="Ej. 50903")
with form_cols[2]:
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0
    st.markdown("&nbsp;", unsafe_allow_html=True)
    if st.button("Limpiar", use_container_width=True):
        st.session_state["uploader_key"] += 1
        st.rerun()

uploaded = None
uploader_cols = st.columns([0.5, 3, 0.5])
with uploader_cols[1]:
    st.markdown("<div class='uploader-label uploader-centered'>Subir XML</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "",
        type=["xml", "zip"],
        accept_multiple_files=True,
        key=f"xml_uploader_{st.session_state['uploader_key']}",
        label_visibility="collapsed",
    )
st.markdown("</div>", unsafe_allow_html=True)

# ===================== Columnas base =====================
base_cols = [
    "Fecha",
    "RFC Emisor",
    "Nombre Emisor",
    "UUID",
    "cfdi:Concepto Importe",
    "cfdi:Traslado Importe",
    "RegimenFiscalReceptor",
]


# ===================== Utilidades CFDI (XML) =====================
def _detect_cfdi_ns(root: ET.Element) -> str:
    if root.tag.startswith("{") and "}Comprobante" in root.tag:
        return root.tag.split("}")[0][1:]
    return "http://www.sat.gob.mx/cfd/4"


def _norm_series(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.upper()


def _norm_text(value: str) -> str:
    return str(value).strip().upper()


def _has_cuenta_predial(xml_bytes: bytes) -> bool:
    root = ET.fromstring(xml_bytes)
    ns = {"cfdi": _detect_cfdi_ns(root)}
    return root.find(".//cfdi:Conceptos/cfdi:Concepto/cfdi:CuentaPredial", ns) is not None


def _get_regimen_emisor(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    ns = {"cfdi": _detect_cfdi_ns(root)}
    emisor = root.find("cfdi:Emisor", ns)
    return (emisor.attrib.get("RegimenFiscal", "") if emisor is not None else "").strip()


def _get_forma_pago(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    return root.attrib.get("FormaPago", "").strip()


def _get_metodo_pago(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    return root.attrib.get("MetodoPago", "").strip()


def _get_tipo_de_comprobante(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    return root.attrib.get("TipoDeComprobante", "").strip()


def _get_domicilio_fiscal_receptor(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    ns = {"cfdi": _detect_cfdi_ns(root)}
    receptor = root.find("cfdi:Receptor", ns)
    return (receptor.attrib.get("DomicilioFiscalReceptor", "") if receptor is not None else "").strip()


def _get_uso_cfdi(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    ns = {"cfdi": _detect_cfdi_ns(root)}
    receptor = root.find("cfdi:Receptor", ns)
    return (receptor.attrib.get("UsoCFDI", "") if receptor is not None else "").strip()


def _get_moneda(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    return root.attrib.get("Moneda", "").strip()


def _get_tipo_cambio(xml_bytes: bytes) -> str:
    root = ET.fromstring(xml_bytes)
    return root.attrib.get("TipoCambio", "").strip()


def _fecha_solo_dia(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return s
    if "T" in s:
        return s.split("T", 1)[0]
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(s[:10], fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            continue
    return s[:10]


def _parse_fecha(value) -> datetime | None:
    try:
        s = _fecha_solo_dia(value)
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None


SINGLE_TC_PATH = Path("data/Tipo Cambio.xls")


def _flatten_cols(cols):
    out = []
    for col in cols:
        if isinstance(col, tuple):
            col = " ".join([str(x) for x in col if str(x) != "None"])
        out.append(str(col).strip())
    return out


def _coerce_tc_date_series(series: pd.Series):
    if np.issubdtype(series.dtype, np.number):
        return pd.to_datetime(series, unit="D", origin="1899-12-30", errors="coerce").dt.date
    out = pd.to_datetime(series.astype(str), errors="coerce", dayfirst=True, infer_datetime_format=True).dt.date
    if pd.isna(out).mean() > 0.5:
        try:
            out2 = pd.to_numeric(series, errors="coerce")
            out_serial = pd.to_datetime(out2, unit="D", origin="1899-12-30", errors="coerce").dt.date
            out = pd.Series(out).where(pd.Series(out).notna(), out_serial).values
        except Exception:
            pass
    return out


def _read_tc_excel_fixed(path: Path) -> pd.Series | None:
    if not path.exists():
        st.error("No se encontró el archivo de tipo de cambio en **data/Tipo Cambio.xls**.")
        return None
    try:
        df = pd.read_excel(path, sheet_name=0, engine="xlrd")
    except Exception:
        try:
            df = pd.read_excel(path, sheet_name=0)
        except Exception as exc:
            st.error(f"No se pudo leer {path}: {exc}")
            return None

    df.columns = _flatten_cols(df.columns)

    tc_col = next((c for c in df.columns if "para solventar obligaciones" in c.lower()), None)
    if tc_col is None:
        st.error("En el Excel de TC no se encontró la columna **'Para solventar obligaciones'**.")
        return None

    fecha_col = next(
        (
            c
            for c in df.columns
            if "fecha" in c.lower() or "public" in c.lower() or np.issubdtype(df[c].dtype, np.datetime64)
        ),
        None,
    )
    if fecha_col is None:
        st.error("En el Excel de TC no se encontró una columna de **fecha**.")
        return None

    work = df[[fecha_col, tc_col]].copy()
    work[fecha_col] = _coerce_tc_date_series(work[fecha_col])
    work[tc_col] = (
        work[tc_col]
        .astype(str)
        .str.replace("\u00a0", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.replace("[^0-9.]", "", regex=True)
    )
    work[tc_col] = pd.to_numeric(work[tc_col], errors="coerce")
    work = work.dropna(subset=[fecha_col, tc_col])

    series = work.drop_duplicates(subset=[fecha_col], keep="last").set_index(fecha_col)[tc_col].sort_index()
    series.name = "TC"
    return series


tc_series = _read_tc_excel_fixed(SINGLE_TC_PATH)


def _tc_lookup(dttm: datetime) -> float | None:
    if tc_series is None or dttm is None:
        return None
    val = tc_series.get(dttm.date())
    return float(val) if pd.notna(val) else None


def _tc_prev_lookup(dttm: datetime) -> float | None:
    if tc_series is None or dttm is None:
        return None
    idx = tc_series.index
    pos = np.searchsorted(idx, dttm.date(), side="left") - 1
    if pos >= 0:
        return float(tc_series.iloc[pos])
    return None


# ===================== Fechas Excel dd/mm/aaaa y escritura segura =====================
def _ensure_excel_date_ddmmyyyy(df: pd.DataFrame, col: str = "Fecha") -> pd.DataFrame:
    if col not in df.columns:
        return df
    out = df.copy()
    s = pd.to_datetime(out[col], errors="coerce", dayfirst=False)
    mask = s.isna()
    if mask.any():
        s2 = pd.to_datetime(out.loc[mask, col], errors="coerce", dayfirst=True)
        s.loc[mask] = s2
    out[col] = s
    return out


def _write_sheet(writer, df: pd.DataFrame, sheet_name: str, title_text: str):
    def _safe_sheet_name(name: str) -> str:
        safe = re.sub(r"[:\\/?*\\[\\]]", "-", str(name)).strip() or "Hoja"
        return safe[:31]

    safe_name = _safe_sheet_name(sheet_name)

    df_x = _ensure_excel_date_ddmmyyyy(df, "Fecha")
    df_x.to_excel(writer, index=False, sheet_name=safe_name, startrow=2)

    wb = writer.book
    ws = wb[safe_name]

    ws["A1"].value = title_text
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    last_col = max(1, df_x.shape[1])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.freeze_panes = "A3"

    last_col_letter = get_column_letter(last_col)
    last_row = df_x.shape[0] + 3
    ws.auto_filter.ref = f"A3:{last_col_letter}{last_row}"

    for i, col in enumerate(df_x.columns, start=1):
        max_len = max([len(str(col))] + [len(str(x)) for x in df_x[col].astype(str).fillna("")])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    if "Fecha" in df_x.columns:
        fecha_idx = list(df_x.columns).index("Fecha") + 1
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=fecha_idx, max_col=fecha_idx):
            row[0].number_format = "DD/MM/YYYY"


try:
    from extractor import parse_cfdi_many
except Exception:
    def parse_cfdi_many(files):
        return []


valid_files: list[tuple[str, bytes]] = []
bad_files: list[str] = []
if uploaded:
    for uf in uploaded:
        name = uf.name or "archivo"
        try:
            data = uf.read()
            if not data:
                raise ValueError("Archivo vacío")

            lower_name = name.lower()
            if lower_name.endswith(".zip"):
                found_xml = False
                try:
                    with ZipFile(io.BytesIO(data)) as zf:
                        for info in zf.infolist():
                            if info.is_dir():
                                continue
                            if not info.filename.lower().endswith(".xml"):
                                continue
                            xml_bytes = zf.read(info)
                            if not xml_bytes:
                                continue
                            ET.fromstring(xml_bytes)
                            valid_files.append((f"{name}:{info.filename}", xml_bytes))
                            found_xml = True
                except BadZipFile as exc:
                    raise ValueError("ZIP inválido") from exc
                if not found_xml:
                    raise ValueError("ZIP sin XML válidos")
            else:
                ET.fromstring(data)
                valid_files.append((name, data))
        except Exception:
            bad_files.append(name)

    if bad_files:
        st.markdown(
            f"<div class='small-error'>⚠️ No se pudo cargar {len(bad_files)} archivo(s): "
            + ", ".join(bad_files[:3]) + ("…" if len(bad_files) > 3 else "") + "</div>",
            unsafe_allow_html=True,
        )

files = valid_files

if files:
    rows = parse_cfdi_many(files)

    for row in rows or []:
        if "Fecha" in row:
            row["Fecha"] = _fecha_solo_dia(row["Fecha"])

    df = pd.DataFrame(rows, columns=base_cols) if rows else pd.DataFrame(columns=base_cols)

    if not df.empty:
        insert_pos = (
            list(df.columns).index("RegimenFiscalReceptor") + 1
            if "RegimenFiscalReceptor" in df.columns
            else len(df.columns)
        )
        df.insert(insert_pos, "Régimen fiscal correcto", regimen_fiscal_correcto)
        mismatches = pd.DataFrame()
        if regimen_fiscal_correcto.strip():
            mismatches = df[
                _norm_series(df["RegimenFiscalReceptor"]) != _norm_series(df["Régimen fiscal correcto"])
            ]

        arrend_rows = []
        for (fname, blob), row in zip(files, rows):
            if _get_regimen_emisor(blob) == "606" and not _has_cuenta_predial(blob):
                arrend_rows.append(row)
        arrend_df = pd.DataFrame(arrend_rows, columns=base_cols) if arrend_rows else pd.DataFrame(columns=base_cols)

        cols_efectivo = [
            "Fecha",
            "RFC Emisor",
            "Nombre Emisor",
            "UUID",
            "cfdi:Concepto Importe",
            "cfdi:Traslado Importe",
            "FormaPago",
        ]
        efectivo_rows = []
        for (fname, blob), row in zip(files, rows):
            forma = _get_forma_pago(blob)
            try:
                total = float(row.get("cfdi:Concepto Importe", 0) or 0) + float(row.get("cfdi:Traslado Importe", 0) or 0)
            except Exception:
                total = 0.0
            if forma == "01" and total > 2000:
                efectivo_rows.append(
                    {
                        "Fecha": row.get("Fecha", ""),
                        "RFC Emisor": row.get("RFC Emisor", ""),
                        "Nombre Emisor": row.get("Nombre Emisor", ""),
                        "UUID": row.get("UUID", ""),
                        "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                        "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                        "FormaPago": forma,
                    }
                )
        efectivo_df = pd.DataFrame(efectivo_rows, columns=cols_efectivo) if efectivo_rows else pd.DataFrame(columns=cols_efectivo)

        cols_domcp = [
            "Fecha",
            "RFC Emisor",
            "Nombre Emisor",
            "UUID",
            "cfdi:Concepto Importe",
            "cfdi:Traslado Importe",
            "DomicilioFiscalReceptor",
            "CP",
        ]
        domcp_rows = []
        if cp_val.strip():
            for (fname, blob), row in zip(files, rows):
                dom_rec = _get_domicilio_fiscal_receptor(blob)
                if _norm_text(dom_rec) != _norm_text(cp_val):
                    domcp_rows.append(
                        {
                            "Fecha": row.get("Fecha", ""),
                            "RFC Emisor": row.get("RFC Emisor", ""),
                            "Nombre Emisor": row.get("Nombre Emisor", ""),
                            "UUID": row.get("UUID", ""),
                            "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                            "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                            "DomicilioFiscalReceptor": dom_rec,
                            "CP": cp_val,
                        }
                    )
        domcp_df = pd.DataFrame(domcp_rows, columns=cols_domcp) if domcp_rows else pd.DataFrame(columns=cols_domcp)

        cols_sin_ret = [
            "Fecha",
            "RFC Emisor",
            "Nombre Emisor",
            "UUID",
            "cfdi:Concepto Importe",
            "cfdi:Traslado Importe",
            "RegimenFiscalEmisor",
            "Retenciones",
        ]
        sin_ret_rows = []
        for (fname, blob), row in zip(files, rows):
            regimen_emisor = _get_regimen_emisor(blob)
            if regimen_emisor != "626":
                continue
            root = ET.fromstring(blob)
            ns = {"cfdi": _detect_cfdi_ns(root)}
            impuestos = set()
            for r in root.findall(".//cfdi:Impuestos/cfdi:Retenciones/cfdi:Retencion", ns):
                imp = r.attrib.get("Impuesto", "").strip()
                if imp:
                    impuestos.add(imp)
            for r in root.findall(".//cfdi:Conceptos/cfdi:Concepto/cfdi:Impuestos/cfdi:Retenciones/cfdi:Retencion", ns):
                imp = r.attrib.get("Impuesto", "").strip()
                if imp:
                    impuestos.add(imp)
            if not any(code in {"001", "002"} for code in impuestos):
                sin_ret_rows.append(
                    {
                        "Fecha": row.get("Fecha", ""),
                        "RFC Emisor": row.get("RFC Emisor", ""),
                        "Nombre Emisor": row.get("Nombre Emisor", ""),
                        "UUID": row.get("UUID", ""),
                        "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                        "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                        "RegimenFiscalEmisor": regimen_emisor,
                        "Retenciones": ", ".join(sorted(impuestos)) if impuestos else "SIN RETENCIONES",
                    }
                )
        sin_ret_df = pd.DataFrame(sin_ret_rows, columns=cols_sin_ret) if sin_ret_rows else pd.DataFrame(columns=cols_sin_ret)

        cols_uso = [
            "Fecha",
            "RFC Emisor",
            "Nombre Emisor",
            "UUID",
            "cfdi:Concepto Importe",
            "cfdi:Traslado Importe",
            "UsoCFDI",
        ]
        uso_rows = []
        for (fname, blob), row in zip(files, rows):
            uso = _get_uso_cfdi(blob).upper()
            if uso == "S01":
                uso_rows.append(
                    {
                        "Fecha": row.get("Fecha", ""),
                        "RFC Emisor": row.get("RFC Emisor", ""),
                        "Nombre Emisor": row.get("Nombre Emisor", ""),
                        "UUID": row.get("UUID", ""),
                        "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                        "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                        "UsoCFDI": uso,
                    }
                )
        uso_df = pd.DataFrame(uso_rows, columns=cols_uso) if uso_rows else pd.DataFrame(columns=cols_uso)

        cols_tc = [
            "Fecha",
            "RFC Emisor",
            "Nombre Emisor",
            "UUID",
            "cfdi:Concepto Importe",
            "cfdi:Traslado Importe",
            "Moneda",
            "TipoCambioXML",
            "TipoCambioDOF",
            "TipoCambioDOFAnterior",
            "Diferencia",
        ]
        tc_rows = []
        for (fname, blob), row in zip(files, rows):
            moneda = _get_moneda(blob).upper()
            if moneda != "USD":
                continue
            fecha_dt = _parse_fecha(row.get("Fecha", ""))
            ref_tc = _tc_lookup(fecha_dt)
            ref_tc_prev = _tc_prev_lookup(fecha_dt)
            xml_tc_raw = _get_tipo_cambio(blob)
            try:
                xml_tc = float(str(xml_tc_raw).replace(",", "")) if xml_tc_raw else None
            except Exception:
                xml_tc = None
            include = False
            if ref_tc is not None and xml_tc is not None:
                include = abs(xml_tc - ref_tc) > 1e-6
            elif ref_tc is not None and xml_tc is None:
                include = True
            if include:
                tc_rows.append(
                    {
                        "Fecha": row.get("Fecha", ""),
                        "RFC Emisor": row.get("RFC Emisor", ""),
                        "Nombre Emisor": row.get("Nombre Emisor", ""),
                        "UUID": row.get("UUID", ""),
                        "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                        "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                        "Moneda": moneda,
                        "TipoCambioXML": xml_tc,
                        "TipoCambioDOF": ref_tc,
                        "TipoCambioDOFAnterior": ref_tc_prev,
                        "Diferencia": (
                            None if (xml_tc is None or ref_tc is None) else round(xml_tc - ref_tc, 6)
                        ),
                    }
                )
        tc_df = pd.DataFrame(tc_rows, columns=cols_tc) if tc_rows else pd.DataFrame(columns=cols_tc)

        usd_all_rows = []
        for (fname, blob), row in zip(files, rows):
            moneda = _get_moneda(blob).upper()
            if moneda != "USD":
                continue
            fecha_dt = _parse_fecha(row.get("Fecha", ""))
            ref_tc = _tc_lookup(fecha_dt)
            ref_tc_prev = _tc_prev_lookup(fecha_dt)
            xml_tc_raw = _get_tipo_cambio(blob)
            try:
                xml_tc = float(str(xml_tc_raw).replace(",", "")) if xml_tc_raw else None
            except Exception:
                xml_tc = None
            dif = None if (xml_tc is None or ref_tc is None) else round(xml_tc - ref_tc, 6)
            usd_all_rows.append(
                {
                    "Fecha": row.get("Fecha", ""),
                    "RFC Emisor": row.get("RFC Emisor", ""),
                    "Nombre Emisor": row.get("Nombre Emisor", ""),
                    "UUID": row.get("UUID", ""),
                    "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                    "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                    "Moneda": moneda,
                    "TipoCambioXML": xml_tc,
                    "TipoCambioDOF": ref_tc,
                    "TipoCambioDOFAnterior": ref_tc_prev,
                    "Diferencia": dif,
                }
            )
        usd_all_df = pd.DataFrame(usd_all_rows, columns=cols_tc) if usd_all_rows else pd.DataFrame(columns=cols_tc)

        cols_nc = [
            "Fecha",
            "RFC Emisor",
            "Nombre Emisor",
            "UUID",
            "cfdi:Concepto Importe",
            "cfdi:Traslado Importe",
            "TipoDeComprobante",
            "FormaPago",
            "MetodoPago",
        ]

        def _norm_forma_pago(fp: str) -> str:
            s = (fp or "").strip()
            s = re.sub(r"^0+", "", s) or "0"
            return s

        nc_rows = []
        for (fname, blob), row in zip(files, rows):
            tipo = _get_tipo_de_comprobante(blob).upper()
            if tipo != "E":
                continue
            forma = _norm_forma_pago(_get_forma_pago(blob))
            metodo = _get_metodo_pago(blob).upper()
            if (forma != "15") or (metodo != "PUE"):
                nc_rows.append(
                    {
                        "Fecha": row.get("Fecha", ""),
                        "RFC Emisor": row.get("RFC Emisor", ""),
                        "Nombre Emisor": row.get("Nombre Emisor", ""),
                        "UUID": row.get("UUID", ""),
                        "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                        "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                        "TipoDeComprobante": tipo,
                        "FormaPago": forma,
                        "MetodoPago": metodo,
                    }
                )
        nc_df = pd.DataFrame(nc_rows, columns=cols_nc) if nc_rows else pd.DataFrame(columns=cols_nc)

        cols_pue99 = [
            "Fecha",
            "RFC Emisor",
            "Nombre Emisor",
            "UUID",
            "cfdi:Concepto Importe",
            "cfdi:Traslado Importe",
            "FormaPago",
            "MetodoPago",
        ]
        pue99_rows = []
        for (fname, blob), row in zip(files, rows):
            tipo = _get_tipo_de_comprobante(blob).upper()
            if tipo != "I":
                continue
            metodo = _get_metodo_pago(blob).upper()
            fp_raw = _get_forma_pago(blob)
            fp_norm = re.sub(r"^0+", "", (fp_raw or "").strip()) or "0"
            if metodo == "PUE" and fp_norm == "99":
                pue99_rows.append(
                    {
                        "Fecha": row.get("Fecha", ""),
                        "RFC Emisor": row.get("RFC Emisor", ""),
                        "Nombre Emisor": row.get("Nombre Emisor", ""),
                        "UUID": row.get("UUID", ""),
                        "cfdi:Concepto Importe": row.get("cfdi:Concepto Importe", 0),
                        "cfdi:Traslado Importe": row.get("cfdi:Traslado Importe", 0),
                        "FormaPago": fp_norm,
                        "MetodoPago": metodo,
                    }
                )
        pue99_df = pd.DataFrame(pue99_rows, columns=cols_pue99) if pue99_rows else pd.DataFrame(columns=cols_pue99)

        dfs = [
            mismatches,
            arrend_df,
            efectivo_df,
            domcp_df,
            sin_ret_df,
            uso_df,
            tc_df,
            usd_all_df,
            nc_df,
            pue99_df,
        ]

        if any(not df_part.empty for df_part in dfs):
            def _make_excel_bytes() -> bytes:
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    _write_sheet(
                        writer,
                        mismatches if not mismatches.empty else pd.DataFrame(columns=df.columns),
                        "No Coinciden",
                        "CFDI con diferente régimen fiscal",
                    )
                    _write_sheet(
                        writer,
                        arrend_df[base_cols] if not arrend_df.empty else pd.DataFrame(columns=base_cols),
                        "Arrendamiento sin Predial",
                        "CFDI de arrendamiento sin cuenta predial",
                    )
                    _write_sheet(
                        writer,
                        efectivo_df if not efectivo_df.empty else pd.DataFrame(columns=cols_efectivo),
                        "Efectivo > 2000",
                        "CFDI con gasto total en efectivo mayor de 2000 pesos",
                    )
                    _write_sheet(
                        writer,
                        domcp_df if not domcp_df.empty else pd.DataFrame(columns=cols_domcp),
                        "Domicilio vs CP",
                        "CFDI con diferente domicilio fiscal",
                    )
                    _write_sheet(
                        writer,
                        sin_ret_df if not sin_ret_df.empty else pd.DataFrame(columns=cols_sin_ret),
                        "Sin Retenciones 001-002 (626)",
                        "Régimen 626 sin retenciones",
                    )
                    _write_sheet(
                        writer,
                        uso_df if not uso_df.empty else pd.DataFrame(columns=cols_uso),
                        "Uso CFDI S01 no deducible",
                        "Uso CFDI S01 no deducible",
                    )
                    _write_sheet(
                        writer,
                        tc_df if not tc_df.empty else pd.DataFrame(columns=cols_tc),
                        "Tipo de Cambio Diferente",
                        "Tipo de cambio diferente",
                    )
                    _write_sheet(
                        writer,
                        usd_all_df if not usd_all_df.empty else pd.DataFrame(columns=cols_tc),
                        "USD (Todos)",
                        "CFDI en USD — Tipo de Cambio DOF",
                    )
                    _write_sheet(
                        writer,
                        nc_df if not nc_df.empty else pd.DataFrame(columns=cols_nc),
                        "NC ≠ Condonación - PUE",
                        "Notas de crédito con método de pago diferente a condonación y PUE",
                    )
                    _write_sheet(
                        writer,
                        pue99_df if not pue99_df.empty else pd.DataFrame(columns=cols_pue99),
                        "PUE FP 99",
                        "PUE con forma de pago por definir",
                    )
                return out.getvalue()

            st.download_button(
                "Descargar Excel",
                _make_excel_bytes(),
                file_name="cfdi_extract.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("No hay filas para exportar en las hojas configuradas.")
    else:
        st.info("No se pudieron leer CFDI válidos. Verifica que los XML estén timbrados y completos.")
