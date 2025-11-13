"""Pantalla del módulo Cédula para actualización y amortización de pérdidas."""

from __future__ import annotations

import io
from copy import copy  # clonar estilos (evitar StyleProxy)
from pathlib import Path
from typing import Dict, List

import pandas as pd
import streamlit as st
from core.theme import apply_theme
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell  # ⬅️ NUEVO: detectar celdas merged
from openpyxl.utils import column_index_from_string, get_column_letter
import re as _re

from urllib.parse import quote

from core.auth import ensure_session_from_token
from core.custom_nav import render_brand_logout_nav
from core.streamlit_compat import set_query_params


ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
NAV_LAYOUT_CSS = """
<style>
  [data-testid="stSidebar"],
  [data-testid="collapsedControl"],
  header[data-testid="stHeader"],
  div[data-testid="stToolbar"],
  #MainMenu,
  #stDecoration,
  #root > div:nth-child(1) > div[data-testid="stSidebarNav"] {
    display: none !important;
  }
  .nav-spacer {
    height: 78px;
  }
</style>
"""


def _auth_query_param() -> str | None:
    raw = st.query_params.get("auth")
    if isinstance(raw, list):
        raw = raw[-1] if raw else None
    if raw:
        return raw
    token = st.session_state.get("auth_token")
    return token if isinstance(token, str) and token else None


def _back_href() -> str:
    href = "?goto=pages/Cedula_Impuestos_inicio.py"
    auth = _auth_query_param()
    if auth:
        href += f"&auth={quote(auth)}"
    return href


def render_fixed_nav() -> None:
    st.markdown(NAV_LAYOUT_CSS, unsafe_allow_html=True)
    render_brand_logout_nav(
        "pages/Cedula_Impuestos_inicio.py",
        brand="Cédulas · Pérdidas",
        action_label="Atrás",
        action_href=_back_href(),
    )
    st.markdown('<div class="nav-spacer"></div>', unsafe_allow_html=True)


def _handle_pending_navigation() -> None:
    params = st.query_params
    raw = params.get("goto")
    goto = None
    if isinstance(raw, list):
        goto = raw[-1] if raw else None
    elif isinstance(raw, str):
        goto = raw
    if goto:
        cleaned = {k: v for k, v in params.items() if k != "goto"}
        set_query_params(cleaned)
        try:
            st.switch_page(goto)
        except Exception:
            pass
        st.stop()

# ─────────────────────────────
# Configuración base
# ─────────────────────────────
YEAR_MIN, YEAR_MAX = 2015, 2024
AMORT_TO = 2025
TEMPLATE_PATH = ASSETS_DIR / "Actualizacion_de_perdidas.xlsx"
OUTPUT_PATH = ASSETS_DIR / "Actualizacion_de_perdidas_actualizado.xlsx"

# Plantilla para “Perdidas declaracion anual.xlsx”
DECL_TEMPLATE_PATH = ASSETS_DIR / "Perdidas declaracion anual.xlsx"
DECL_OUTPUT_PATH = ASSETS_DIR / "Perdidas_declaracion_anual_actualizado.xlsx"

# Límites para evitar procesos largos
MAX_EXTRA_BLOCKS = 8  # máximo de copias J11:O37→...
MAX_STAGES = 3 + MAX_EXTRA_BLOCKS  # bloques totales visibles (1:C, 2:I, 3:O ya existen)


# ─────────────────────────────
# Utilidades para Excel
# ─────────────────────────────
def _col_to_idx(col: str) -> int:
    return column_index_from_string(col)


def _idx_to_col(idx: int) -> str:
    return get_column_letter(idx)


def _iter_range(a1: str, b1: str) -> List[str]:
    ma = _re.match(r"([A-Z]+)(\d+)", a1)
    mb = _re.match(r"([A-Z]+)(\d+)", b1)
    if not ma or not mb:
        return []
    ac, ar = ma.groups()
    bc, br = mb.groups()
    ia, ib = _col_to_idx(ac), _col_to_idx(bc)
    ra, rb = int(ar), int(br)
    coords = []
    for ci in range(min(ia, ib), max(ia, ib) + 1):
        for rr in range(min(ra, rb), max(ra, rb) + 1):
            coords.append(f"{_idx_to_col(ci)}{rr}")
    return coords


# ─────────────────────────────
# Evaluador de fórmulas (compacto)
# ─────────────────────────────
def _evaluate_cell(ws, coord: str, memo: Dict[str, float]) -> float:
    if coord in memo:
        return memo[coord]

    val = ws[coord].value

    if isinstance(val, (int, float)):
        memo[coord] = float(val)
        return memo[coord]
    if val is None or (isinstance(val, str) and val.strip() == ""):
        memo[coord] = 0.0
        return 0.0

    if isinstance(val, str) and val.startswith("="):
        expr = val[1:].lstrip("+").replace("$", "")

        def replace_refs(text: str) -> str:
            refs = _re.findall(r"\b[A-Z]{1,3}\d{1,7}\b", text)
            for r in sorted(set(refs), key=len, reverse=True):
                text = _re.sub(rf"\b{r}\b", str(_evaluate_cell(ws, r, memo)), text)
            return text

        def repl_round_match(m):
            inner = replace_refs(m.group(2))
            decs = int(m.group(3))
            try:
                v = eval(inner, {"__builtins__": {}}, {})
            except Exception:
                v = 0.0
            return str(round(v, decs))

        expr = _re.sub(
            r"(?i)\b(FIXED|ROUND|REDONDEAR)\(\s*([^,]+)\s*,\s*(-?\d+)\s*\)",
            repl_round_match,
            expr,
        )

        def repl_if(m):
            cond, tval, fval = m.group(2), m.group(3), m.group(4)
            cond = replace_refs(cond).replace("<>", "!=")
            cond = _re.sub(r"(?<![<>!])=(?!=)", "==", cond)
            try:
                ok = bool(eval(cond, {"__builtins__": {}}, {}))
            except Exception:
                ok = False
            chosen = replace_refs(tval if ok else fval)
            try:
                return str(eval(chosen, {"__builtins__": {}}, {}))
            except Exception:
                try:
                    return str(float(chosen))
                except Exception:
                    return "0"

        expr = _re.sub(
            r"(?i)\b(IF|SI)\(\s*([^,]+)\s*,\s*([^,]+)\s*,\s*([^)]+)\)",
            repl_if,
            expr,
        )

        def repl_sum(m):
            inside = m.group(2)
            parts, depth, buff = [], 0, ""
            for ch in inside:
                if ch == "," and depth == 0:
                    parts.append(buff)
                    buff = ""
                else:
                    if ch == "(":
                        depth += 1
                    elif ch == ")":
                        depth -= 1
                    buff += ch
            if buff:
                parts.append(buff)
            total = 0.0
            for p in parts:
                p = p.strip()
                rng = _re.match(r"\b([A-Z]{1,3}\d{1,7}):([A-Z]{1,3}\d{1,7})\b", p)
                if rng:
                    for coord2 in _iter_range(rng.group(1), rng.group(2)):
                        total += _evaluate_cell(ws, coord2, memo)
                else:
                    p = replace_refs(p)
                    try:
                        total += float(eval(p, {"__builtins__": {}}, {}))
                    except Exception:
                        total += 0.0
            return str(total)

        expr = _re.sub(r"(?i)\b(SUM|SUMA)\(\s*([^)]+)\s*\)", repl_sum, expr)

        def repl_minmax(m):
            fn = m.group(1).upper()
            inside = replace_refs(m.group(2))
            try:
                vals = [float(x) for x in inside.split(",")]
            except Exception:
                vals = [0.0]
            return str((min if fn == "MIN" else max)(vals))

        expr = _re.sub(r"(?i)\b(MIN|MAX)\(\s*([^)]+)\s*\)", repl_minmax, expr)

        expr = replace_refs(expr)
        try:
            res = eval(expr, {"__builtins__": {}}, {})
        except Exception:
            res = 0.0

        memo[coord] = float(res)
        return memo[coord]

    memo[coord] = 0.0
    return 0.0


def compute_B3_to_L3(wb) -> List[float]:
    ws = wb.active
    memo: Dict[str, float] = {}
    cols = list("BCDEFGHIJKL")
    return [_evaluate_cell(ws, f"{c}3", memo) for c in cols]


def write_inputs_to_sheet(wb, rows: List[dict]):
    ws = wb.active
    base_row = 33  # N33=2015
    for row in rows:
        year = int(row["AÑO"])
        if not (YEAR_MIN <= year <= YEAR_MAX):
            continue
        idx = base_row + (year - YEAR_MIN)
        ws.cell(idx, 14, year)  # N
        ws.cell(idx, 15, float(row.get("PERDIDA") or 0))  # O
        util = row.get("UTILIDAD")
        if util not in (None, ""):
            ws.cell(idx, 16, float(util))  # P
    return wb


def lookup_amort_value_above(wb, loss_year: int, amort_year: int) -> float:
    ws = wb.active
    row_idx = None
    for r in range(3, 200):
        v = ws[f"N{r}"].value
        if isinstance(v, (int, float)) and int(v) == int(loss_year):
            row_idx = r
            break
    if row_idx is None:
        return 0.0
    col_letter = None
    for c in [chr(x) for x in range(ord("O"), ord("Y") + 1)]:
        v = ws[f"{c}2"].value
        if isinstance(v, (int, float)) and int(v) == int(amort_year):
            col_letter = c
            break
    if col_letter is None:
        return 0.0
    coord = f"{col_letter}{max(1, row_idx - 1)}"
    return _evaluate_cell(ws, coord, memo={})


def write_amortizations_to_sheet(wb, rows: pd.DataFrame):
    ws = wb.active
    for _, r in rows.iterrows():
        loss_year = int(r["Año (Pérdida)"])
        amort_year = int(r["Año (Amortización)"])
        amount = float(r.get("Amortización") or 0.0)
        row_idx = None
        for rr in range(3, 200):
            v = ws[f"N{rr}"].value
            if isinstance(v, (int, float)) and int(v) == loss_year:
                row_idx = rr
                break
        if row_idx is None:
            continue
        col_letter = None
        for c in [chr(x) for x in range(ord("O"), ord("Y") + 1)]:
            v = ws[f"{c}2"].value
            if isinstance(v, (int, float)) and int(v) == amort_year:
                col_letter = c
                break
        if col_letter is None:
            continue
        ws[f"{col_letter}{row_idx}"].value = amount
    return wb


# ─────────────────────────────
# Orden y recálculo incremental
# ─────────────────────────────
def sort_rows(df: pd.DataFrame) -> pd.DataFrame:
    return df.sort_values(
        ["Año (Pérdida)", "Año (Amortización)", "SEQ"], kind="mergesort"
    ).reset_index(drop=True)


def recalc_rows_incremental(base_wb_bytes: bytes, df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = sort_rows(df.copy())
    for loss_year in df["Año (Pérdida)"].unique():
        mask = df["Año (Pérdida)"] == loss_year
        grp_idx = df.index[mask].tolist()
        wb = load_workbook(io.BytesIO(base_wb_bytes), data_only=False)
        for idx in grp_idx:
            ay = int(df.at[idx, "Año (Amortización)"])
            value_above = float(lookup_amort_value_above(wb, int(loss_year), ay))
            df.at[idx, "Perdida Actualizada del Año"] = value_above
            amt = float(df.at[idx, "Amortización"] or 0.0)
            df.at[idx, "Remanente de Pérdida"] = value_above - amt
            wb = write_amortizations_to_sheet(
                wb,
                pd.DataFrame(
                    [
                        {
                            "Año (Pérdida)": int(loss_year),
                            "Año (Amortización)": ay,
                            "Amortización": amt,
                        }
                    ]
                ),
            )
    return df


# ─────────────────────────────
# Tabla 1 / Tabla 3 helpers
# ─────────────────────────────
def find_table1_col_by_year(ws, year: int) -> int | None:
    for c in range(_col_to_idx("A"), _col_to_idx("L") + 1):
        v = ws[f"{_idx_to_col(c)}2"].value
        if isinstance(v, (int, float)) and int(v) == int(year):
            return c
    return None


def table1_value(ws, row_idx: int, year: int) -> float:
    col = find_table1_col_by_year(ws, year)
    if col is None:
        return 0.0
    return _evaluate_cell(ws, f"{_idx_to_col(col)}{row_idx}", memo={})


def table3_loss_amount(ws, loss_year: int) -> float:
    row_idx = 33 + (int(loss_year) - 2015)
    return _evaluate_cell(ws, f"O{row_idx}", memo={})


# ─────────────────────────────
# Copia de bloque J11:O37 con estilos/merges  (limitada)
# ─────────────────────────────
ORD_MAP = {
    1: "PRIMERA",
    2: "SEGUNDA",
    3: "TERCERA",
    4: "CUARTA",
    5: "QUINTA",
    6: "SEXTA",
    7: "SEPTIMA",
    8: "OCTAVA",
    9: "NOVENA",
    10: "DECIMA",
}


def _copy_cell_style(src, dst):
    """Clonar estilos para evitar 'unhashable type: StyleProxy'."""

    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def copy_block_with_style(ws, src_tl: str, src_br: str, dst_tl: str):
    ma = _re.match(r"([A-Z]+)(\d+)", src_tl)
    mb = _re.match(r"([A-Z]+)(\d+)", src_br)
    md = _re.match(r"([A-Z]+)(\d+)", dst_tl)
    sc, sr = _col_to_idx(ma.group(1)), int(ma.group(2))
    ec, er = _col_to_idx(mb.group(1)), int(mb.group(2))
    dc, dr = _col_to_idx(md.group(1)), int(md.group(2))
    ncols, nrows = ec - sc + 1, er - sr + 1

    for r in range(nrows):
        for c in range(ncols):
            src_cell = ws.cell(sr + r, sc + c)
            dst_cell = ws.cell(dr + r, dc + c)
            dst_cell.value = src_cell.value
            _copy_cell_style(src_cell, dst_cell)

    new_ranges = []
    for mr in list(ws.merged_cells.ranges):
        tl, br = str(mr).split(":")
        ta, tb = _re.match(r"([A-Z]+)(\d+)", tl), _re.match(r"([A-Z]+)(\d+)", br)
        mc, mrw = _col_to_idx(ta.group(1)), int(ta.group(2))
        nc, nrw = _col_to_idx(tb.group(1)), int(tb.group(2))
        if sc <= mc <= ec and sc <= nc <= ec and sr <= mrw <= er and sr <= nrw <= er:
            off_c, off_r = dc - sc, dr - sr
            new_tl = f"{_idx_to_col(mc + off_c)}{mrw + off_r}"
            new_br = f"{_idx_to_col(nc + off_c)}{nrw + off_r}"
            new_ranges.append(f"{new_tl}:{new_br}")
    for rng in new_ranges:
        if rng not in [str(x) for x in ws.merged_cells.ranges]:
            ws.merge_cells(rng)


def last_year_in_row13(ws, start_col_letter="J", width=6) -> tuple[int, int]:
    row_year = 13
    last_idx = _col_to_idx("O")
    last_year = ws[f"{_idx_to_col(last_idx)}{row_year}"].value
    last_year = int(last_year) if isinstance(last_year, (int, float)) else None
    while True:
        next_last_idx = last_idx + width
        v = ws[f"{_idx_to_col(next_last_idx)}{row_year}"].value
        if isinstance(v, (int, float)):
            last_idx = next_last_idx
            last_year = int(v)
        else:
            break
    return last_year if last_year is not None else 0, last_idx


def ensure_blocks_until_year(ws, target_year: int):
    """
    Copia **J11:O37** hacia la derecha y sube el año de la última col (fila 13)
    hasta alcanzar `target_year`, pero con un **tope de MAX_EXTRA_BLOCKS** para
    evitar procesos largos.
    """

    src_tl, src_br = "J11", "O37"
    width = 6
    row_year = 13

    cur_year, last_idx = last_year_in_row13(ws, "J", width)
    block_ord = 4  # el primer bloque nuevo es la CUARTA actualización
    made = 0

    while cur_year < int(target_year) and made < MAX_EXTRA_BLOCKS:
        dest_start_col = last_idx + 1  # p.ej., O→P
        dst_tl = f"{_idx_to_col(dest_start_col)}11"
        copy_block_with_style(ws, src_tl, src_br, dst_tl)

        # Título del nuevo bloque
        for r in range(11, 38):
            for c in range(dest_start_col, dest_start_col + width):
                val = ws.cell(r, c).value
                if isinstance(val, str) and "ACTUALIZACIÓN" in val.upper():
                    ws.cell(r, c).value = f"{ORD_MAP.get(block_ord, '')} ACTUALIZACIÓN"
                    break

        # Año en fila 13 (última col del bloque)
        cur_year += 1
        ws[f"{_idx_to_col(dest_start_col + width - 1)}{row_year}"].value = int(cur_year)

        last_idx = dest_start_col + width - 1
        block_ord += 1
        made += 1


# ─────────────────────────────
# Llenado por BLOQUES desde Tabla 1 + marcadores de fila 13
# ─────────────────────────────
def stage_data_col_idx(stage: int) -> int:
    """Columna base del bloque: 1→C, 2→I, 3→O, 4→U, 5→AA, 6→AG, …"""

    return _col_to_idx("C") + (stage - 1) * 6


def rows_for_stage(stage: int) -> tuple[int, int, int] | None:
    """
    stage 1: (C19,C27,C31,C34,C37) se maneja aparte.
    stage >=2: start = 12 + 7*(stage-2); devuelve (start, start+2, start+5)
    """

    if stage == 1:
        return None
    start = 12 + 7 * (stage - 2)
    return (start, start + 2, start + 5)


def mark_years_row13(ws_decl, start_year: int, end_year: int, max_stages: int = MAX_STAGES):
    """
    Escribe C13=start_year, I13, O13, U13… incrementando de 1 en 1,
    limitado a `max_stages`.
    """

    year = int(start_year)
    stage = 1
    while year <= int(end_year) and stage <= max_stages:
        col_idx = stage_data_col_idx(stage)
        ws_decl[f"{_idx_to_col(col_idx)}13"].value = year
        stage += 1
        year += 1


def clear_unused_blocks(ws_decl, keep_stages: int):
    """
    Limpia el contenido de los bloques a la derecha de `keep_stages`
    para que no aparezcan bloques (Octava, Novena, etc.) cuando no se requieren.
    No altera estilos/merges; **no** intenta escribir sobre celdas merged no-ancla.
    """

    for stage in range(keep_stages + 1, MAX_STAGES + 1):
        c0 = stage_data_col_idx(stage)
        c1 = c0 + 5
        for r in range(11, 38):  # filas del bloque J11:O37
            for c in range(c0, c1 + 1):
                cell = ws_decl.cell(r, c)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None


def fill_stage_from_table1(ws_decl, ws_src, loss_year: int, stage: int):
    """Llena los datos del bloque `stage` desde la Tabla 1."""

    if stage == 1:
        ws_decl["C19"].value = table1_value(ws_src, 6, loss_year)
        ws_decl["C27"].value = table1_value(ws_src, 7, loss_year)
        ws_decl["C31"].value = table1_value(ws_src, 8, loss_year)
        ws_decl["C34"].value = table3_loss_amount(ws_src, loss_year)
        ws_decl["C37"].value = table1_value(ws_src, 10, loss_year)
    else:
        r = rows_for_stage(stage)
        if not r:
            return
        base_col = _idx_to_col(stage_data_col_idx(stage))
        ws_decl[f"{base_col}19"].value = table1_value(ws_src, r[0], loss_year)
        ws_decl[f"{base_col}31"].value = table1_value(ws_src, r[1], loss_year)
        ws_decl[f"{base_col}37"].value = table1_value(ws_src, r[2], loss_year)


def fill_all_stages_from_table1(ws_decl, ws_src, loss_year: int, target_year: int):
    """
    Asegura bloques hasta `target_year` (con tope), marca C13/I13/O13/…,
    llena bloques necesarios y **borra** los que sobren para que no se muestren.
    """

    ensure_blocks_until_year(ws_decl, target_year)
    total_stages = min(MAX_STAGES, int(target_year) - int(loss_year) + 1)
    end_year = int(loss_year) + total_stages - 1
    mark_years_row13(ws_decl, loss_year, end_year, max_stages=MAX_STAGES)
    for stg in range(1, total_stages + 1):
        fill_stage_from_table1(ws_decl, ws_src, loss_year, stg)
    clear_unused_blocks(ws_decl, keep_stages=total_stages)


# ─────────────────────────────
# Declaración: armado final
# ─────────────────────────────
def _find_row_col_for(ws, loss_year: int, amort_year: int):
    row_idx = None
    for r in range(3, 300):
        v = ws[f"N{r}"].value
        if isinstance(v, (int, float)) and int(v) == int(loss_year):
            row_idx = r
            break
    if row_idx is None:
        return None, None, None
    col_letter = None
    for c in [chr(x) for x in range(ord("O"), ord("Y") + 1)]:
        v = ws[f"{c}2"].value
        if isinstance(v, (int, float)) and int(v) == int(amort_year):
            col_letter = c
            break
    if col_letter is None:
        return None, None, None
    return row_idx, _col_to_idx(col_letter), col_letter


def fill_declaracion_excel(
    declaracion_year: int,
    utilidad_digitada: float,
    base_bytes_without_amorts: bytes,
    amort_rows_df: pd.DataFrame,
) -> io.BytesIO:
    if not DECL_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"No se encontró la plantilla de declaración: {DECL_TEMPLATE_PATH.name}")

    wb_state = load_workbook(io.BytesIO(base_bytes_without_amorts), data_only=False)
    if amort_rows_df is not None and not amort_rows_df.empty:
        wb_state = write_amortizations_to_sheet(wb_state, amort_rows_df)
    ws_src = wb_state.active

    if amort_rows_df is None or amort_rows_df.empty:
        raise ValueError("No hay filas de amortización/pérdida para la declaración.")
    loss_years = sorted(set(int(y) for y in amort_rows_df["Año (Pérdida)"].tolist()))
    if not loss_years:
        raise ValueError("No se detectaron años con pérdida.")
    first_loss_year = loss_years[0]

    row_idx, col_idx, col_letter = _find_row_col_for(ws_src, first_loss_year, declaracion_year)
    if row_idx is None:
        raise ValueError("No se encontró cruce en Cuadro 2 (verifica años en N y fila 2).")
    memo = {}
    up_left_col_idx = max(1, col_idx - 1)
    up_left_col_letter = _idx_to_col(up_left_col_idx)
    val_G9 = _evaluate_cell(ws_src, f"{up_left_col_letter}{max(1, row_idx - 1)}", memo)
    val_I9 = _evaluate_cell(ws_src, f"{col_letter}{max(1, row_idx - 1)}", memo)
    val_K9 = min(val_I9, float(utilidad_digitada))
    val_M9 = max(val_I9 - float(utilidad_digitada), 0.0)

    decl_wb = load_workbook(DECL_TEMPLATE_PATH, data_only=False)
    decl_ws = decl_wb.active

    decl_ws["M1"].value = int(declaracion_year)
    decl_ws["C3"].value = float(utilidad_digitada)
    decl_ws["E9"].value = int(first_loss_year)
    decl_ws["G9"].value = float(val_G9)
    decl_ws["I9"].value = float(val_I9)
    decl_ws["K9"].value = float(val_K9)
    decl_ws["M9"].value = float(val_M9)

    fill_all_stages_from_table1(decl_ws, ws_src, first_loss_year, int(declaracion_year))

    out = io.BytesIO()
    decl_wb.save(out)
    out.seek(0)
    return out


def main() -> None:
    """Renderiza la pantalla protegida dentro del módulo Cédula."""

    ensure_session_from_token()
    st.set_page_config(page_title="Actualización y Amortización de Pérdidas", layout="wide")
apply_theme()
    _handle_pending_navigation()
    render_fixed_nav()

    st.title("Actualización y Amortización de Pérdidas")
    st.caption(
        "Captura pérdidas por año. Al subir, se escribe en la **Tabla 3 (N32:P43)** de la plantilla fija "
        "`assets/Actualizacion_de_perdidas.xlsx`, se calculan **B3:L3** y se muestra un cuadro de resultados "
        "solo con años que tuvieron pérdida. En ese cuadro eliges el **Año (Amortización)**; "
        "la **Pérdida Actualizada del Año** se toma del **Cuadro 2 (N2:Y24)** buscando el año en col. **N** "
        "y el año de amortización en **fila 2**, y tomando **la celda de arriba** del cruce."
    )

    # ─────────────────────────────
    # UI: Parámetros de declaración
    # ─────────────────────────────
    st.subheader("0) Parámetros de declaración (para llenar 'Perdidas declaracion anual.xlsx')")
    cA, cB, cC = st.columns([1, 1, 2])
    with cA:
        declaracion_year = st.number_input(
            "Año de declaración → va a **M1**",
            min_value=YEAR_MIN + 1,
            max_value=AMORT_TO,
            value=AMORT_TO,
            step=1,
            help="Año que corresponde a la declaración (se escribe en M1 del Excel de declaración).",
            key="decl_year",
        )
    with cB:
        utilidad_digitada = st.number_input(
            "Utilidad digitada → va a **C3**",
            min_value=0.0,
            step=1.0,
            value=0.0,
            help="Monto de utilidad que se coloca en C3 del Excel de declaración.",
            key="util_digit",
        )
    with cC:
        if not DECL_TEMPLATE_PATH.exists():
            st.error(f"⚠️ No se encontró `{DECL_TEMPLATE_PATH.name}` en assets/.")
        else:
            st.info(f"Plantilla de declaración detectada: `{DECL_TEMPLATE_PATH.name}`")

    # ─────────────────────────────
    # UI: Captura de datos base
    # ─────────────────────────────
    with st.container():
        st.subheader("1) Captura de datos")
        if not TEMPLATE_PATH.exists():
            st.error(
                f"No se encontró la plantilla en: `{TEMPLATE_PATH}`.\n\n"
                "👉 Asegúrate de **commitear** el archivo y que `.gitignore` NO lo excluya."
            )
            st.stop()
        st.info(f"Plantilla en uso (repositorio): `{TEMPLATE_PATH.name}`")

    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(
            [{"AÑO": YEAR_MIN, "PERDIDA": 0.0, "UTILIDAD": 0.0, "AMORTIZACIÓN": 0.0}]
        )

    colconfig = {
        "AÑO": st.column_config.SelectboxColumn(
            options=list(range(YEAR_MIN, YEAR_MAX + 1)), required=True, width="small"
        ),
        "PERDIDA": st.column_config.NumberColumn(format="%.2f", min_value=0.0, step=0.01),
        "UTILIDAD": st.column_config.NumberColumn(format="%.2f", step=0.01),
        "AMORTIZACIÓN": st.column_config.NumberColumn(
            format="%.2f",
            step=0.01,
            help="No se escribe en Tabla 3",
        ),
    }

    c1, c2, c3 = st.columns([1, 2, 1])
    with c1:
        used_years = set(int(y) for y in st.session_state.df["AÑO"].tolist() if pd.notna(y))
        available_years = [y for y in range(YEAR_MIN, YEAR_MAX + 1) if y not in used_years]
        if st.button("➕ Agregar año", disabled=(len(available_years) == 0), use_container_width=True):
            next_year = available_years[0] if available_years else YEAR_MIN
            st.session_state.df = pd.concat(
                [
                    st.session_state.df,
                    pd.DataFrame(
                        [
                            {
                                "AÑO": next_year,
                                "PERDIDA": 0.0,
                                "UTILIDAD": 0.0,
                                "AMORTIZACIÓN": 0.0,
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )

    st.session_state.df = st.data_editor(
        st.session_state.df,
        column_config=colconfig,
        use_container_width=True,
        num_rows="dynamic",
        key="editor_perdidas",
    )

    # ─────────────────────────────
    # Acción: subir y calcular
    # ─────────────────────────────
    st.subheader("2) Subir al Excel y calcular B3:L3")
    b1, b2, b3 = st.columns(3)
    pressed = b2.button("⬆️ Subir información", use_container_width=True)

    if pressed:
        df_in = st.session_state.df.sort_index().groupby("AÑO", as_index=False, sort=True).last()
        wb = load_workbook(TEMPLATE_PATH, data_only=False)
        wb = write_inputs_to_sheet(wb, df_in.to_dict(orient="records"))
        b3_l3_vals = compute_B3_to_L3(wb)
        years_2015_2025 = list(range(2015, 2026))
        updated_map = {y: b3_l3_vals[i] for i, y in enumerate(years_2015_2025)}
        df_losses = df_in[df_in["PERDIDA"].fillna(0) > 0].copy().sort_values("AÑO")

        rows = []
        seq = 0
        for _, r in df_losses.iterrows():
            y = int(r["AÑO"])
            seq += 1
            rows.append(
                {
                    "SEQ": seq,
                    "Eliminar": False,
                    "Año (Pérdida)": y,
                    "Perdida Actualizada": float(updated_map.get(y, 0.0)),
                    "Año (Amortización)": max(y + 1, YEAR_MIN + 1),
                    "Perdida Actualizada del Año": 0.0,
                    "Amortización": 0.0,
                    "Remanente de Pérdida": 0.0,
                }
            )

        st.session_state._wb_base_bytes = io.BytesIO()
        wb.save(st.session_state._wb_base_bytes)
        st.session_state._wb_base_bytes.seek(0)
        base_bytes = st.session_state._wb_base_bytes.getvalue()
        st.session_state._updated_map = updated_map
        st.session_state.result_rows = recalc_rows_incremental(base_bytes, pd.DataFrame(rows))
        st.session_state.next_seq = seq + 1

    # ─────────────────────────────
    # Resultados + descarga
    # ─────────────────────────────
    if "result_rows" in st.session_state:
        st.subheader("3) Resultados (sólo años con pérdida)")

        with st.expander("➕ Agregar amortización (una por vez)", expanded=False):
            loss_years_opts = sorted(set(st.session_state.result_rows["Año (Pérdida)"].tolist()))
            ly = st.selectbox("Año (Pérdida)", options=loss_years_opts, key="add_loss_y")
            am_opts = list(range(int(ly) + 1, AMORT_TO + 1))
            ay = st.selectbox("Año (Amortización)", options=am_opts, key="add_amort_y")
            amt = st.number_input("Amortización", min_value=0.0, step=1.0, key="add_amt")
            if st.button("Agregar fila"):
                new_seq = int(st.session_state.next_seq)
                new_row = {
                    "SEQ": new_seq,
                    "Eliminar": False,
                    "Año (Pérdida)": int(ly),
                    "Perdida Actualizada": float(st.session_state._updated_map.get(int(ly), 0.0)),
                    "Año (Amortización)": int(ay),
                    "Perdida Actualizada del Año": 0.0,
                    "Amortización": float(amt),
                    "Remanente de Pérdida": 0.0,
                }
                st.session_state.result_rows = pd.concat(
                    [st.session_state.result_rows, pd.DataFrame([new_row])], ignore_index=True
                )
                st.session_state.next_seq = new_seq + 1
                base_bytes = st.session_state._wb_base_bytes.getvalue()
                st.session_state.result_rows = recalc_rows_incremental(base_bytes, st.session_state.result_rows)

        amort_options = list(range(YEAR_MIN + 1, AMORT_TO + 1))
        colconfig2 = {
            "Eliminar": st.column_config.CheckboxColumn(help="Marca y elimina"),
            "Año (Pérdida)": st.column_config.NumberColumn(format="%d", disabled=True, width="small"),
            "Perdida Actualizada": st.column_config.NumberColumn(format="%.2f", disabled=True),
            "Año (Amortización)": st.column_config.SelectboxColumn(
                options=amort_options, required=True, width="small"
            ),
            "Perdida Actualizada del Año": st.column_config.NumberColumn(format="%.2f", disabled=True),
            "Amortización": st.column_config.NumberColumn(format="%.2f", step=1.0),
            "Remanente de Pérdida": st.column_config.NumberColumn(format="%.2f", disabled=True),
        }

        df_edit = st.data_editor(
            sort_rows(st.session_state.result_rows),
            column_config=colconfig2,
            use_container_width=True,
            num_rows="fixed",
            key="editor_resultados",
        )

        if st.button("🗑️ Eliminar seleccionados", use_container_width=False):
            df_edit = df_edit[~df_edit["Eliminar"]].copy()

        base_bytes = st.session_state._wb_base_bytes.getvalue()
        df_edit = recalc_rows_incremental(base_bytes, df_edit)

        st.session_state.result_rows = df_edit
        st.dataframe(df_edit.drop(columns=["SEQ", "Eliminar"]), use_container_width=True)

        wb2 = load_workbook(io.BytesIO(base_bytes), data_only=False)
        wb2 = write_amortizations_to_sheet(wb2, df_edit)
        bio = io.BytesIO()
        wb2.save(bio)
        bio.seek(0)
        st.download_button(
            "💾 Descargar Excel actualizado",
            data=bio,
            file_name="Actualizacion_de_perdidas_actualizado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.subheader("4) Generar archivo de declaración (bloques C/I/O/U/AA… + C13→… hasta M1)")
        gen = st.button("🧾 Llenar y descargar 'Perdidas declaracion anual.xlsx'", use_container_width=True)
        if gen:
            st.session_state.pop("decl_bytes", None)
            if not DECL_TEMPLATE_PATH.exists():
                st.error(f"No se encontró `{DECL_TEMPLATE_PATH.name}` en assets/.")
            else:
                try:
                    out_bytes = fill_declaracion_excel(
                        declaracion_year=int(st.session_state.get("decl_year", AMORT_TO)),
                        utilidad_digitada=float(st.session_state.get("util_digit", 0.0)),
                        base_bytes_without_amorts=base_bytes,
                        amort_rows_df=st.session_state.result_rows,
                    )
                    st.session_state.decl_bytes = out_bytes.getvalue()
                    st.success("Archivo de declaración generado.")
                except Exception as e:
                    st.error(f"No se pudo generar el archivo de declaración: {e!s}")

        if "decl_bytes" in st.session_state:
            st.download_button(
                "⬇️ Descargar Declaración",
                data=st.session_state.decl_bytes,
                file_name="Perdidas_declaracion_anual_actualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_decl",
            )

    st.markdown("---")
    st.caption(
        f"El copiado de **J11:O37→P11** está limitado a **{MAX_EXTRA_BLOCKS}** bloques extra "
        f"(total visible hasta **{MAX_STAGES}** etapas). Se escriben los años en **C13, I13, O13, U13, AA13, …** "
        "comenzando con el año de la pérdida y avanzando de uno en uno hasta alcanzar **M1**. "
        "Los bloques sobrantes a la derecha se **limpian** sin escribir sobre celdas mergeadas (evita errores)."
    )


if __name__ == "__main__":
    main()
