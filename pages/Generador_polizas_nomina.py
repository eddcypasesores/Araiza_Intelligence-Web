"""Generador de pólizas contables para provisión de nómina."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from urllib.parse import urlencode

import streamlit as st
from openpyxl import Workbook, load_workbook
import re
import unicodedata

from core.theme import apply_theme
from core.auth import ensure_session_from_token, auth_query_params
from core.custom_nav import handle_logout_request, render_brand_logout_nav


def _get_params() -> dict[str, str]:
    try:
        raw = st.query_params
    except Exception:
        raw = st.experimental_get_query_params()

    flattened: dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(value, list):
            value = value[-1] if value else None
        if value is None:
            continue
        flattened[key] = str(value)
    return flattened


def _handle_pending_navigation() -> None:
    params = _get_params()
    goto = params.pop("goto", None)
    if not goto:
        return
    try:
        st.query_params.clear()
        if params:
            st.query_params.update(params)
    except Exception:
        st.experimental_set_query_params(**params)
    try:
        st.switch_page(goto)
    except Exception:
        st.stop()
    st.stop()


def _back_href() -> str:
    params = {"goto": "pages/generador_polizas.py"}
    params.update(auth_query_params())
    return f"?{urlencode(params, doseq=False)}"


st.set_page_config(
    page_title="Generador Provisión Nómina",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed",
)
apply_theme()
ensure_session_from_token()
handle_logout_request()
_handle_pending_navigation()

PAGE_CSS = """
<style>
#MainMenu,
header[data-testid="stHeader"],
footer,
div[data-testid="stToolbar"],
[data-testid="stSidebar"],
[data-testid="collapsedControl"] {
  display:none !important;
}
.block-container {
  padding-top: 120px !important;
  max-width: 900px !important;
}
body, [data-testid="stAppViewContainer"] {
  background:#f5f6fb !important;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

render_brand_logout_nav(
    "pages/generador_polizas.py",
    brand="Generador de pólizas",
    action_label="Atrás",
    action_href=_back_href(),
)


def obtener_dia(valor_fecha):
    if isinstance(valor_fecha, (datetime, date)):
        return valor_fecha.day
    if isinstance(valor_fecha, str):
        parsed = None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                parsed = datetime.strptime(valor_fecha, fmt)
                break
            except Exception:
                continue
        if parsed is not None:
            return parsed.day
        return valor_fecha
    return valor_fecha


def es_diferente_cero(val):
    if val is None:
        return False
    try:
        return float(str(val).replace(",", ".")) != 0.0
    except Exception:
        return True


def a_numero(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    try:
        return float(str(val))
    except Exception:
        return 0


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _resolve_column(ws, default_letter: str | None, keywords: list[str]) -> str | None:
    header_row = ws[1]
    for cell in header_row:
        header = _normalize_header(cell.value)
        if header and all(keyword in header for keyword in keywords):
            return cell.column_letter
    return default_letter


def _flag_marked(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return abs(float(value)) > 0
    text = str(value).strip().lower()
    if not text:
        return False
    if text in {"si", "sí", "true", "x", "1"}:
        return True
    try:
        return float(text.replace(",", ".")) != 0.0
    except Exception:
        return False


def generar_provision(wb_src, consecutivo_inicial: int):
    ws_src = wb_src.active

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Provisión Nómina"

    next_row = 3
    consecutivo = int(consecutivo_inicial)
    max_row = ws_src.max_row

    col_v = _resolve_column(ws_src, "V", ["RETEN", "ISR"])
    col_w = _resolve_column(ws_src, None, ["ISR", "ASIMIL"])
    col_isr_aguinaldo = _resolve_column(ws_src, None, ["ISR", "AGUINAL"])
    col_x = _resolve_column(ws_src, "X", ["IMSS"])
    col_y = _resolve_column(ws_src, "Y", ["INFONAVIT"])
    col_subsidio_flag = _resolve_column(ws_src, None, ["SUBSIDIO", "CAUSADO"])
    col_prima_dominical = _resolve_column(ws_src, None, ["PRIMA", "DOMINICAL"])

    for src_row in range(2, max_row + 1):
        k = ws_src[f"K{src_row}"].value
        r = ws_src[f"R{src_row}"].value
        s = ws_src[f"S{src_row}"].value
        t = ws_src[f"T{src_row}"].value
        u = ws_src[f"U{src_row}"].value
        ret_isr_val = a_numero(ws_src[f"{col_v}{src_row}"].value) if col_v else 0
        isr_asimilados_val = (
            a_numero(ws_src[f"{col_w}{src_row}"].value) if col_w else 0
        )
        isr_aguinaldo_val = (
            a_numero(ws_src[f"{col_isr_aguinaldo}{src_row}"].value) if col_isr_aguinaldo else 0
        )
        x = ws_src[f"{col_x}{src_row}"].value
        y = ws_src[f"{col_y}{src_row}"].value
        z = ws_src[f"Z{src_row}"].value
        aa = ws_src[f"AA{src_row}"].value
        prima_dominical_val = (
            ws_src[f"{col_prima_dominical}{src_row}"].value if col_prima_dominical else 0
        )
        subsidio_flag_value = (
            ws_src[f"{col_subsidio_flag}{src_row}"].value if col_subsidio_flag else None
        )
        has_subsidio_flag = _flag_marked(subsidio_flag_value)

        m = ws_src[f"M{src_row}"].value
        n = ws_src[f"N{src_row}"].value
        o = ws_src[f"O{src_row}"].value
        importe_prov = a_numero(m) + a_numero(o) - a_numero(n)

        if not any(
            [
                es_diferente_cero(r),
                es_diferente_cero(s),
                es_diferente_cero(t),
                es_diferente_cero(u),
                es_diferente_cero(ret_isr_val),
                es_diferente_cero(isr_asimilados_val),
                es_diferente_cero(isr_aguinaldo_val),
                es_diferente_cero(x),
                es_diferente_cero(y),
                es_diferente_cero(z),
                es_diferente_cero(aa),
                es_diferente_cero(prima_dominical_val),
            ]
        ):
            continue

        header_row = next_row
        ws_new[f"A{header_row}"] = "Dr"
        ws_new[f"B{header_row}"] = consecutivo
        ws_new[f"C{header_row}"] = "Provisión Nomina"
        ws_new[f"D{header_row}"] = obtener_dia(k)

        next_row = header_row + 1

        def _cargo_entry(code: str, descripcion: str, monto):
            nonlocal next_row
            ws_new[f"B{next_row}"] = code
            ws_new[f"C{next_row}"] = 0
            ws_new[f"D{next_row}"] = descripcion
            ws_new[f"E{next_row}"] = 1
            ws_new[f"F{next_row}"] = monto if monto is not None else 0
            ws_new[f"G{next_row}"] = 0
            next_row += 1

        def _abono_entry(code: str, descripcion: str, monto):
            nonlocal next_row
            ws_new[f"B{next_row}"] = code
            ws_new[f"C{next_row}"] = 0
            ws_new[f"D{next_row}"] = descripcion
            ws_new[f"E{next_row}"] = 1
            ws_new[f"F{next_row}"] = 0
            ws_new[f"G{next_row}"] = monto if monto is not None else 0
            next_row += 1

        if es_diferente_cero(r):
            _cargo_entry("4102-002-000001", "Sueldos y Salarios", r)
        if es_diferente_cero(s):
            _cargo_entry("4102-002-000013", "Sueldo Asimilado a Salario", s)
        if es_diferente_cero(t):
            _cargo_entry("4102-002-000006", "Aguinaldo", t)
        if es_diferente_cero(u):
            _cargo_entry("4102-002-000012", "Comisiones", u)
        if es_diferente_cero(prima_dominical_val):
            _cargo_entry("4102-002-000012", "Prima Dominical", prima_dominical_val)
        if has_subsidio_flag and es_diferente_cero(z):
            _cargo_entry("1108-001-000000", "Subsidio para el Empleo", z)
        isr_sueldos_val = max(ret_isr_val - isr_asimilados_val, 0.0)
        if es_diferente_cero(isr_sueldos_val):
            _abono_entry("2103-001-000000", "ISR Sueldos y Salario", isr_sueldos_val)
        if es_diferente_cero(isr_aguinaldo_val):
            _abono_entry("2103-001-000000", "ISR Aguinaldo", isr_aguinaldo_val)
        if es_diferente_cero(isr_asimilados_val):
            _abono_entry("2103-008-000000", "ISR x Asimilados a Salario", isr_asimilados_val)
        if es_diferente_cero(x):
            _abono_entry("2103-002-000000", "IMSS Trabajador", x)
        if es_diferente_cero(y):
            _abono_entry("2103-006-000000", "Credito Infonavit", y)
        if has_subsidio_flag and es_diferente_cero(aa):
            _abono_entry("1108-001-000000", "Subsidio para el Empleo", aa)
        if es_diferente_cero(r):
            _abono_entry("2105-001-000000", "Provisión de sueldos y salarios por pagar", importe_prov)
        if es_diferente_cero(s):
            _abono_entry("2105-002-000000", "Provisión de asimilados a salarios x pagar", importe_prov)

        ws_new[f"B{next_row}"] = "FIN_PARTIDAS"
        next_row += 1
        consecutivo += 1

    return wb_new


st.title("Generador de Provisión Nómina")

b3_val = st.number_input(
    "Valor inicial para el consecutivo (columna B del encabezado)",
    value=1,
    step=1,
    min_value=0,
)

st.write("Sube el archivo de nómina en Excel (ejemplo: `nomina_xml.xlsx`).")
uploaded_file = st.file_uploader("Archivo de nómina", type=["xlsx"])

if uploaded_file is None:
    st.info("Sube un archivo para comenzar.")
else:
    st.success("Archivo cargado correctamente.")
    if st.button("Generar archivo de Provisión Nómina"):
        try:
            wb_src = load_workbook(uploaded_file, data_only=True)
            wb_new = generar_provision(wb_src, b3_val)
        except Exception as exc:
            st.error(f"No se pudo generar la provisión: {exc}")
        else:
            buffer = BytesIO()
            wb_new.save(buffer)
            buffer.seek(0)
            st.download_button(
                label="Descargar provision_nomina.xlsx",
                data=buffer.getvalue(),
                file_name="provision_nomina.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _resolve_column(ws, default_letter: str, keywords: list[str]) -> str:
    header_row = ws[1]
    for cell in header_row:
        header = _normalize_header(cell.value)
        if header and all(keyword in header for keyword in keywords):
            return cell.column_letter
    return default_letter
